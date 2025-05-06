import os
import subprocess
import sys
import shutil
import re

# Importar bibliotecas de GUI
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk

# Importar bibliotecas de OCR e conversão de PDF
from pdf2image import convert_from_path
import pytesseract

# Importar bibliotecas de manipulação de planilhas
import openpyxl
from openpyxl import Workbook

# Importar bibliotecas para abrir URLs e multithreading
import webbrowser
from threading import Thread



# Função para verificar e instalar pacotes automaticamente
def install_package(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Função para baixar e configurar o Poppler automaticamente
def install_poppler():
    # Verifica se o Poppler está instalado
    try:
        subprocess.check_call(["poppler-utils", "--version"])
        print("Poppler já está instalado.")
        return
    except FileNotFoundError:
        print("Poppler não encontrado. Instalando...")

    # Instala o Poppler usando pip
    try:
        install_package("poppler-utils")
        print("Poppler instalado com sucesso.")
    except Exception as e:
        print(f"Erro ao instalar o Poppler: {e}")

    # Define o caminho do Poppler para o sistema
def get_poppler_path():
    if getattr(sys, 'frozen', False):  # Verifica se está rodando como executável
        # Caminho relativo no executável
        base_path = sys._MEIPASS
        return os.path.join(base_path, 'poppler/bin')
    else:
        # Caminho usado durante o desenvolvimento
        return r'C:\poppler-utils\bin'

# Definir o caminho do Poppler
poppler_path = get_poppler_path()
os.environ["PATH"] = poppler_path + os.sep + os.environ["PATH"]


# Chamar a função para instalar o Poppler
install_poppler()

# Definir o caminho do Tesseract OCR
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Função para centralizar as janelas
def center_window(window):
    window.update_idletasks()  # Atualiza a geometria da janela antes de pegar as dimensões
    width = window.winfo_width()
    height = window.winfo_height()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    
    # Calcula a posição x e y para centralizar a janela
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    
    window.geometry(f'{width}x{height}+{x}+{y}')

# Função para fazer OCR em PDFs
def ocr_pdf(pdf_path, output_dir):
    try:
        images = convert_from_path(pdf_path)
        text = ""
        for image in images:
            text += pytesseract.image_to_string(image, lang='por')  # OCR em português
        return text
    except Exception as e:
        print(f"Erro ao fazer OCR em {pdf_path}: {e}")
        return None

# Função para extrair informações do texto OCR
def extract_info(text):
    # Exibir o texto extraído para depuração
    print("Texto extraído:")
    print(text)

    # Regex para CNPJ (flexível para capturar variações como "CNPJ" sem os dois pontos)
    cnpj_match = re.search(r'CNPJ\s*:\s*(\d{2}\.\d{3}\.\d{3}\/\d{4}\-\d{2})', text)
    cnpj = cnpj_match.group(1) if cnpj_match else 'N/A'

    # Caracteres que NÃO devem estar na linha digitável
    invalid_chars = r'[\/\\\(\)\[\]\{\}:]'

    # Encontrar linha com mais de 30 caracteres numéricos, ignorando espaços e pontos
    linha_digitavel = 'N/A'
    for line in text.splitlines():
        cleaned_line = re.sub(r'[^0-9]', '', line)  # Remove todos os caracteres que não são números
        # Verifica se a linha limpa tem mais de 30 caracteres e se a linha original não contém caracteres inválidos
        if len(cleaned_line) > 30 and not re.search(invalid_chars, line):
            linha_digitavel = line  # Preserva a linha original (com pontos ou espaços)
            break

    # Extrair valor do boleto (exemplo simples; ajuste conforme necessário)
    valor_monetario = 'N/A'
    if linha_digitavel != 'N/A':
        # Tenta usar os últimos 10 dígitos "limpos" para formar o valor monetário
        cleaned_digits = re.sub(r'[^0-9]', '', linha_digitavel)  # Remove caracteres não numéricos
        if len(cleaned_digits) >= 10:
            valor_monetario = f"{cleaned_digits[-10:-2]},{cleaned_digits[-2:]}"  # Formato de moeda

    return {
        'cnpj': cnpj,
        'linha_digitavel': linha_digitavel,
        'valor': valor_monetario
    }

# Função para processar PDFs e gerar a planilha Excel
def process_pdfs(input_dir, output_dir, result_file):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Cria uma nova planilha ou carrega uma existente
    if not os.path.exists(result_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Boletos"
        ws.append(['Nome do Arquivo', 'CNPJ Beneficiado', 'Linha Digitável', 'Valor do Boleto'])
    else:
        wb = openpyxl.load_workbook(result_file)
        ws = wb.active

    processed_files = []  # Lista para controlar arquivos processados

    for file_name in os.listdir(input_dir):
        if file_name.endswith('.pdf') and file_name not in processed_files:
            pdf_path = os.path.join(input_dir, file_name)

            # Atualiza o label da janela pop-up
            popup_file_label.config(text=f"Processando: {file_name}")
            root.update_idletasks()

            # Fazer OCR no PDF
            ocr_text = ocr_pdf(pdf_path, output_dir)
            if ocr_text:
                # Extrair informações
                info = extract_info(ocr_text)

                # Adicionar informações à planilha
                ws.append([file_name[:-4], info['cnpj'], info['linha_digitavel'], info['valor']])

            processed_files.append(file_name)  # Adiciona o arquivo à lista de processados

            # Atualiza a barra de progresso
            popup_progress_bar['value'] += 1
            root.update_idletasks()

    # Formatar coluna "Valor do Boleto" como moeda
    for cell in ws['D']:
        cell.number_format = 'R$ #,##0.00'

    # Salvar planilha
    wb.save(result_file)
    print(f"Planilha salva em {result_file}")

def select_input_dir():
    global input_dir
    input_dir = filedialog.askdirectory(title="Selecione a pasta de entrada")
    input_dir_label.config(text=f"Pasta de entrada: {input_dir}")

def select_result_file():
    global result_file
    result_file = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Planilhas Excel", "*.xlsx")],
        title="Salve a planilha de resultados",
    )
    result_file_label.config(text=f"Arquivo de resultados: {result_file}")

def process_pdfs_thread(input_dir, output_dir, result_file):
    global processing, popup_progress_bar
    processing = True
    try:
        # Move os PDFs para uma pasta temporária para evitar conflitos
        temp_dir = os.path.join(input_dir, 'temp')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)

        pdf_files = []  # Lista para armazenar os nomes dos arquivos PDF
        for file_name in os.listdir(input_dir):
            if file_name.endswith('.pdf'):
                source_path = os.path.join(input_dir, file_name)
                destination_path = os.path.join(temp_dir, file_name)
                shutil.move(source_path, destination_path)
                pdf_files.append(file_name)

        # Atualiza a barra de progresso
        popup_progress_bar['value'] = 0
        popup_progress_bar['maximum'] = len(pdf_files)

        # Processa os PDFs
        process_pdfs(temp_dir, temp_dir, result_file)

        if processing:  # Verifica se o processamento foi concluído normalmente
            messagebox.showinfo("Sucesso", "Processamento concluído!")

            # Move os PDFs de volta para a pasta de entrada
            for file_name in pdf_files:
                source_path = os.path.join(temp_dir, file_name)
                destination_path = os.path.join(input_dir, file_name)
                shutil.move(source_path, destination_path)

            # Remove a pasta temporária
            shutil.rmtree(temp_dir)

            # Abre a planilha gerada
            webbrowser.open_new_tab(f"file://{result_file}")  # Abre a planilha no navegador
        else:
            messagebox.showinfo("Informação", "Processamento cancelado pelo usuário.")

            # Remove a pasta temporária (se foi criada)
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro durante o processamento: {e}")
    finally:
        processing = False
        popup.destroy()

def start_processing():
    global processing, popup, popup_file_label, popup_progress_bar
    processing = True  # Flag para indicar se o processamento está ativo

    # Verifica se as pastas e arquivos foram selecionados
    if not input_dir or not result_file:
        messagebox.showwarning("Aviso", "Por favor, selecione a pasta de entrada e o arquivo de resultados.")
        return

    # Janela pop-up
    popup = tk.Toplevel(root)
    popup.title("Processando PDFs")
    
    # Layout do pop-up
    popup_file_label = tk.Label(popup, text="Processando arquivo...", font=("Arial", 12))
    popup_file_label.pack(pady=10)

    popup_progress_bar = ttk.Progressbar(popup, orient="horizontal", length=300, mode="determinate")
    popup_progress_bar.pack(pady=10)

    cancel_button = tk.Button(popup, text="Cancelar", command=cancel_processing)
    cancel_button.pack(pady=10)

    # Centraliza a janela pop-up
    center_window(popup)

    # Iniciar a thread para processar os PDFs
    thread = Thread(target=process_pdfs_thread, args=(input_dir, input_dir, result_file))
    thread.start()

def cancel_processing():
    global processing
    processing = False  # Define a flag para False, indicando que o processamento foi cancelado

# Janela principal
root = tk.Tk()
root.title("Boletos 2 Excel Filter for Alechander ❤")

# Layout
frame = tk.Frame(root)
frame.pack(padx=20, pady=20)

input_dir_button = tk.Button(frame, text="Selecionar Pasta de Entrada", command=select_input_dir)
input_dir_button.grid(row=0, column=0, padx=5, pady=5)

input_dir_label = tk.Label(frame, text="Pasta de entrada não selecionada")
input_dir_label.grid(row=0, column=1, padx=5, pady=5)

result_file_button = tk.Button(frame, text="Salvar Planilha de Resultados", command=select_result_file)
result_file_button.grid(row=1, column=0, padx=5, pady=5)

result_file_label = tk.Label(frame, text="Arquivo de resultados não selecionado")
result_file_label.grid(row=1, column=1, padx=5, pady=5)

start_button = tk.Button(frame, text="Iniciar Processamento", command=start_processing)
start_button.grid(row=2, columnspan=2, pady=10)

# Centraliza a janela principal
center_window(root)

root.mainloop()
