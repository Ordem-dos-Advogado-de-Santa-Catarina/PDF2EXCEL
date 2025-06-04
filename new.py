import os
import subprocess
import sys
import re
from tkinter import *
from tkinter import filedialog, messagebox, ttk, font, scrolledtext # Importar scrolledtext e ttk
from pdf2image import convert_from_path
import pytesseract
import openpyxl
from openpyxl import Workbook
import webbrowser
from threading import Thread
import logging
import tkinter as tk
import glob
from openpyxl.styles import Alignment, PatternFill, Font
import csv
import tempfile
import shutil

# *** Configura√ß√£o de Logging ***
log_dir = os.path.join(os.environ['APPDATA'], 'PDF2EXCEL')
if not os.path.exists(log_dir):
    os.makedirs(log_dir)
log_file_path = os.path.join(log_dir, 'PDF2EXCEL.log')

# Configura√ß√£o do logger global para ser manipulado pela GUI
logger = logging.getLogger()
logger.setLevel(logging.INFO)
# Limpa handlers existentes para evitar duplica√ß√£o em re-execu√ß√µes (√∫til para desenvolvimento)
if logger.handlers:
    for handler in logger.handlers[:]: # Itera sobre uma c√≥pia para permitir modifica√ß√£o
        logger.removeHandler(handler)
        if isinstance(handler, logging.FileHandler):
            handler.close() # Garante que o arquivo seja fechado

# Adiciona handler de arquivo
file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logger.addHandler(file_handler)

logger.info("Programa PDF2EXCEL iniciado.") # Primeiro log ao iniciar

# *** Configura√ß√£o do Filtro de CNPJ ***
filtro_config_path = os.path.join(log_dir, 'Filtro.config')
ignored_cnpjs_list = []
DEFAULT_IGNORED_CNPJ = "82.519.190/0001-12" # CNPJ da OAB como padr√£o

def load_ignored_cnpjs():
    global ignored_cnpjs_list
    try:
        if os.path.exists(filtro_config_path):
            with open(filtro_config_path, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if content:
                    ignored_cnpjs_list = [cnpj.strip() for cnpj in content.split(',') if cnpj.strip()]
                else:
                    # Se o arquivo existe mas est√° vazio, usa o padr√£o e salva
                    ignored_cnpjs_list = [DEFAULT_IGNORED_CNPJ]
                    save_cnpjs_to_config(DEFAULT_IGNORED_CNPJ) # Salva o padr√£o se o arquivo estiver vazio
        else:
            # Se o arquivo n√£o existe, cria com o padr√£o
            ignored_cnpjs_list = [DEFAULT_IGNORED_CNPJ]
            save_cnpjs_to_config(DEFAULT_IGNORED_CNPJ) # Salva o padr√£o se o arquivo n√£o existir
    except Exception as e:
        logger.error(f"Erro ao carregar CNPJs ignorados: {e}")
        ignored_cnpjs_list = [DEFAULT_IGNORED_CNPJ] # Fallback para o padr√£o em caso de erro
    logger.info(f"CNPJs ignorados carregados: {ignored_cnpjs_list}")

def save_cnpjs_to_config(cnpjs_string):
    global ignored_cnpjs_list
    try:
        with open(filtro_config_path, 'w', encoding='utf-8') as f:
            f.write(cnpjs_string)
        # Recarrega a lista ap√≥s salvar
        load_ignored_cnpjs()
        logger.info(f"CNPJs ignorados salvos em {filtro_config_path}: {cnpjs_string}")
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar CNPJs no arquivo de configura√ß√£o: {e}")
        messagebox.showerror("Erro", f"N√£o foi poss√≠vel salvar o arquivo de configura√ß√£o de CNPJs: {e}", icon="error")
        return False

# Carrega os CNPJs ignorados na inicializa√ß√£o
load_ignored_cnpjs()

# Definir o caminho do Poppler (apenas para a vers√£o .exe)
poppler_path = os.path.join(sys._MEIPASS, 'poppler', 'bin') if getattr(sys, 'frozen', False) else r"C:\Program Files\poppler\bin"

# Se o Poppler n√£o estiver na pasta padr√£o do execut√°vel, verifica se est√° instalado em C:\Program Files\poppler\bin
if getattr(sys, 'frozen', False) and not os.path.exists(poppler_path):
    poppler_path = r"C:\Program Files\poppler\bin"  # Define o caminho alternativo

# Definir o caminho do Tesseract OCR (apenas para a vers√£o .exe)
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Vari√°vel global para controle de processamento (usada para cancelar a thread)
processing = False

# Centraliza o programa na tela
def center_window(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    window.geometry(f'+{x}+{y}')

def ocr_pdf(pdf_path, temp_dir):
    try:
        images = convert_from_path(pdf_path, poppler_path=poppler_path, output_folder=temp_dir, paths_only=False, fmt='jpeg')
        num_pages_in_pdf = len(images)
        text = ""
        for image_idx, image in enumerate(images):
            if not processing: # Checa a flag global de cancelamento
                logger.info(f"Processamento OCR de {os.path.basename(pdf_path)} cancelado na p√°gina {image_idx+1}.")
                break
            text += pytesseract.image_to_string(image, lang='por')
            # N√£o atualiza a barra de progresso por p√°gina OCRizada aqui.
            # A barra de progresso principal agora avan√ßa por PDF processado.
            pass
        return text, num_pages_in_pdf
    except Exception as e:
        logger.exception(f"Erro ao processar OCR do PDF: {pdf_path}")
        return None, 0

def extract_info(text):
    global ignored_cnpjs_list
    cnpj = None
    linhas_digitaveis = []
    valores_monetarios = []
    numero_guia = None
    valor = None

    cnpj_matches = re.findall(r'(\d{2}\.\d{3}\.\d{3}\/\d{4}\-\d{2})', text)
    valid_cnpjs = [cnp for cnp in cnpj_matches if cnp not in ignored_cnpjs_list]
    if valid_cnpjs:
        cnpj = valid_cnpjs[0]
    else:
        cnpj = 'N/A'

    if "GUIA √öNICA DE CUSTAS" in text:
        numero_guia_match = re.search(r"N¬∫ da Guia\s*([\d\.]+/\d+)", text)
        if numero_guia_match:
            numero_guia = numero_guia_match.group(1)
        valor_match = re.search(r"R\$\s*([\d,.]+)", text)
        if valor_match:
            valor = valor_match.group(1)
        return {
            'cnpj': cnpj,
            'numero_guia': numero_guia,
            'valor': valor,
            'linhas_digitaveis': [],
            'valores': [],
            'tipo': 'guia_custas'
        }
    else:
        text = re.sub(r'\d{3}-\d', '', text)
        for line in text.splitlines():
            cleaned_line = re.sub(r'[^0-9]', '', line)
            if 47 <= len(cleaned_line) <= 48:
                linhas_digitaveis.append(cleaned_line)
                valor_monetario = f"{cleaned_line[-10:-2]},{cleaned_line[-2:]}"
                valores_monetarios.append(valor_monetario)
        return {
            'cnpj': cnpj,
            'linhas_digitaveis': linhas_digitaveis,
            'valores': valores_monetarios,
            'numero_guia': None,
            'valor': None,
            'tipo': 'boleto'
        }

# Fun√ß√µes globais auxiliares que ainda podem ser chamadas.
def create_rounded_button(parent, text, command, width=20, height=20):
    canvas = Canvas(parent, width=width, height=height, bd=0, highlightthickness=0, relief='ridge', bg=parent.cget("bg"))
    canvas.create_oval(1, 1, width-2, height-2, outline="#0000FF", fill="#0000FF") # Cor azul para o "i"
    canvas.create_text(width/2, height/2, text=text, fill="#FFFFFF", font=("Segoe UI Bold", int(height/2)))
    canvas.bind("<Button-1>", lambda event: command())
    return canvas

class AppCelescReporter:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("Extrator de Dados Do Boleto (1.4.0a)")
        font_style = font.Font(family="Segoe UI Bold", size=15)
        self.root.option_add("*Font", font_style)
        self.root.resizable(False, False)

        # Vari√°veis de controle de progresso
        self.total_pages_to_process = 0 # Total de PDFs a processar (para a barra principal)
        self.processed_pages_count = 0  # Contador de PDFs processados (para a barra principal)
        self.input_files = [] # Inicializa input_files da inst√¢ncia
        self.result_file = "" # Inicializa result_file da inst√¢ncia
        self.input_dir = "" # Inicializa input_dir da inst√¢ncia

        # Estilo para a barra de progresso
        style = ttk.Style(self.root)
        style.theme_use('clam')
        style.configure("Default.Horizontal.TProgressbar", troughcolor='white', background='green')
        style.configure("Success.Horizontal.TProgressbar", troughcolor='white', background='green')
        style.configure("Error.Horizontal.TProgressbar", troughcolor='white', background='red')

        # Estilo para os bot√µes Iniciar e Cancelar (agora para o bot√£o principal)
        style.configure("Process.TButton", font=("Segoe UI Bold", 15), background="#4CAF50", foreground="white")
        style.map("Process.TButton", background=[('active', '#4CAF50')])
        style.configure("Cancel.TButton", font=("Segoe UI Bold", 15), background="#FF0000", foreground="white") # Cor vermelha para "Cancelar"
        style.map("Cancel.TButton", background=[('active', '#FF0000')])


        main_frame = ttk.Frame(self.root, padding="20 20 20 20")
        main_frame.pack(expand=True, fill=tk.BOTH)

        # --- Se√ß√£o de Sele√ß√£o de Arquivos ---
        input_frame = ttk.LabelFrame(main_frame, text="Configura√ß√µes", padding="10")
        input_frame.pack(fill=tk.X, pady=5)

        # Selecionar PDFs
        input_dir_button = ttk.Button(input_frame, text=" Selecionar PDFs ", command=self.select_input_files)
        input_dir_button.grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.input_dir_label = ttk.Label(input_frame, text="PDF n√£o selecionado", width=45, anchor="w")
        self.input_dir_label.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # Local do Resultado
        result_file_button = ttk.Button(input_frame, text="Local do Resultado", command=self.select_result_file)
        result_file_button.grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.result_file_label = ttk.Label(input_frame, text="Arquivo n√£o selecionado", width=45, anchor="w")
        self.result_file_label.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # CSV ponto e v√≠rgula
        self.save_csv_var = tk.BooleanVar()
        save_csv_label = ttk.Label(input_frame, text="CSV ponto e v√≠rgula:", cursor="hand2")
        save_csv_label.grid(row=2, column=0, pady=5, sticky="e")
        save_csv_label.bind("<Button-1>", lambda event: self.toggle_csv_save())
        self.save_csv_check = ttk.Checkbutton(input_frame, variable=self.save_csv_var)
        self.save_csv_check.grid(row=2, column=1, pady=5, sticky="w")

        # Campo Custas
        custas_label = ttk.Label(input_frame, text="Custas:", anchor='e', justify='left')
        custas_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.custas_entry = ttk.Entry(input_frame, width=10)
        self.custas_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        self.custas_entry.bind("<KeyRelease>", self.limit_custas_entry)

        # --- Se√ß√£o Log em Tempo Real ---
        log_frame = ttk.LabelFrame(main_frame, text="Log de Processamento", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=8, state=tk.DISABLED, font=("Segoe UI", 9), bg="white", fg="black")
        self.log_text.pack(fill=tk.BOTH, expand=True)
        # Configura√ß√£o das tags de cor para o log
        self.log_text.tag_config("INFO", foreground="black")
        self.log_text.tag_config("WARNING", foreground="orange")
        self.log_text.tag_config("ERROR", foreground="red")
        self.log_text.tag_config("CRITICAL_ERROR", foreground="darkred", font=('TkDefaultFont', 9, 'bold'))
        self.log_text.tag_config("SUCCESS", foreground="green")
        self.log_text.tag_config("DEBUG", foreground="gray")

        # Adiciona um handler para redirecionar logs do Python para o widget
        self.log_handler = TextLogHandler(self.log_text)
        logger.addHandler(self.log_handler)

        # --- Se√ß√£o de A√ß√£o (Barra de Progresso e Bot√µes de Controle) ---
        action_frame = ttk.Frame(main_frame, padding="10")
        action_frame.pack(fill=tk.X, pady=10)

        self.progress_bar = ttk.Progressbar(action_frame, orient="horizontal", length=300, mode="determinate",
                                            style="Default.Horizontal.TProgressbar")
        self.progress_bar.pack(pady=5, fill=tk.X)

        self.status_label = ttk.Label(action_frame, text="Aguardando configura√ß√£o...", font=("Segoe UI", 9, "italic"))
        self.status_label.pack(fill=tk.X, pady=5)

        button_row_frame = ttk.Frame(action_frame)
        button_row_frame.pack(pady=5)

        # O bot√£o principal que alterna entre "Iniciar" e "Cancelar"
        self.process_button = ttk.Button(button_row_frame, text="             Iniciar Processamento             ", command=self.start_processing, style="Process.TButton")
        self.process_button.pack(side=tk.LEFT, padx=5) # Centraliza melhor se for o √∫nico bot√£o aqui

        # Bot√£o de Informa√ß√£o "i"
        self.show_info_button_canvas = create_rounded_button(self.root, "i", self.show_info, width=20, height=20)
        self.show_info_button_canvas.place(relx=1.0, rely=1.0, x=-10, y=-10, anchor="se")

        center_window(self.root)
        self.root.bind("<Return>", lambda event: self.start_processing() if not self.root.grab_current() else "break")

    def limit_custas_entry(self, event):
        current_text = self.custas_entry.get()
        if len(current_text) > 5:
            new_text = current_text[:5]
            if self.custas_entry.get() != new_text:
                self.custas_entry.delete(0, tk.END)
                self.custas_entry.insert(0, new_text)

    def set_progress_bar_style(self, style_name):
        """Define o estilo visual da barra de progresso."""
        self.progress_bar.config(style=style_name)

    def log_message(self, message, level="INFO"):
        """Adiciona uma mensagem ao log (thread-safe via after) e tamb√©m ao arquivo de log."""
        if level == "INFO": logger.info(message)
        elif level == "WARNING": logger.warning(message)
        elif level == "ERROR": logger.error(message)
        elif level == "CRITICAL_ERROR": logger.critical(message)
        elif level == "DEBUG": logger.debug(message)
        elif level == "SUCCESS": logger.info(message) # SUCCESS √© tratado como INFO no arquivo de log

    def update_progress_ocr_page(self, pages_processed):
        """Este m√©todo n√£o √© mais usado para a barra de progresso principal,
           mas poderia ser usado para um sub-progresso ou detalhe no status label.
           Por simplicidade, a barra principal agora avan√ßa por PDF.
        """
        # Mantido por compatibilidade, mas n√£o usado para barra principal.
        # A barra de progresso principal √© atualizada por PDF em _actual_processing_task.
        # O status_label ser√° atualizado com o nome do PDF sendo processado.
        pass

    def _update_main_button_state(self, state):
        """Atualiza o texto, comando e estilo do bot√£o principal de processamento."""
        global processing # Acessa a flag global para sincroniza√ß√£o

        if state == 'initial': # Estado inicial ou ap√≥s conclus√£o/cancelamento
            self.process_button.config(text="             Iniciar Processamento             ",
                                       command=self.start_processing,
                                       style="Process.TButton",
                                       state=tk.NORMAL)
            processing = False # Garante que a flag global esteja False
            self.status_label.config(text="Aguardando configura√ß√£o...")
            self.progress_bar.config(value=0)
            self.set_progress_bar_style("Default.Horizontal.TProgressbar")
            self.processed_pages_count = 0 # Resetar contador de PDFs processados

        elif state == 'processing_start': # Quando o processamento come√ßa
            self.process_button.config(text="             Cancelar Processamento             ",
                                       command=self.cancel_processing_gui,
                                       style="Cancel.TButton", # Usar o estilo de cancelamento
                                       state=tk.NORMAL) # Bot√£o habilitado para permitir o cancelamento
            processing = True # Define a flag global para iniciar

        elif state == 'cancelling_pending': # Ap√≥s clicar em cancelar, antes da thread terminar
            self.process_button.config(text="             Cancelando... Aguarde             ",
                                       command=self.cancel_processing_gui, # A√ß√£o ainda √© cancelar, mas desabilitado
                                       style="Cancel.TButton",
                                       state=tk.DISABLED) # Desabilita para evitar m√∫ltiplos cliques
            self.status_label.config(text="Cancelando... Por favor, aguarde.")
            self.set_progress_bar_style("Error.Horizontal.TProgressbar")
            processing = False # Sinaliza o cancelamento na flag global

    def _actual_processing_task(self, input_files_list, output_dir_str, result_file_str, custas_str, save_csv_bool):
        """Cont√©m o loop principal de processamento de PDF, executa em uma thread separada."""
        global processing # Acessa a vari√°vel global 'processing'

        self.log_message("Thread de processamento iniciada.", "DEBUG")

        self.total_pages_to_process = len(input_files_list) # Maximo da barra √© o n√∫mero de PDFs
        self.root.after(0, lambda: self.progress_bar.config(maximum=self.total_pages_to_process))
        self.root.after(0, lambda: self.status_label.config(text=f"Iniciando processamento de {self.total_pages_to_process} documentos..."))
        self.root.after(0, lambda: self.set_progress_bar_style("Default.Horizontal.TProgressbar"))

        error_messages = []
        arquivos_com_paginas_a_mais = set()
        arquivos_com_dados_incompletos = set()
        linhas_digitaveis_processadas = set() # Inicializado por execu√ß√£o da thread

        temp_dir_obj = tempfile.TemporaryDirectory()
        temp_dir = temp_dir_obj.name
        self.log_message(f"Pasta tempor√°ria criada para OCR: {temp_dir}", "DEBUG")

        try:
            wb = None
            ws = None

            if not os.path.exists(result_file_str):
                wb = Workbook()
                ws = wb.active
                ws.title = "Boletos"
                ws.append(['Obeserva√ß√£o', 'Fornecedor', 'C√≥digo de Barras', 'Valor', 'Nome do Titulo'])
                ws.freeze_panes = 'A2' # Congela a primeira linha (cabe√ßalho)
                self.log_message("Novo arquivo Excel criado para resultados.", "INFO")
            else:
                try:
                    wb = openpyxl.load_workbook(result_file_str)
                    ws = wb.active
                    ws.freeze_panes = 'A2' # Garante o congelamento da primeira linha ao abrir
                    self.log_message(f"Arquivo Excel existente carregado: {result_file_str}", "INFO")
                except Exception as e:
                    self.log_message(f"Erro ao abrir arquivo Excel existente: {result_file_str} - {e}", "ERROR")
                    messagebox.showerror("Erro", f"N√£o foi poss√≠vel abrir o arquivo Excel: {e}", icon="error")
                    processing = False # Define a flag global para parar
                    return # Sai da thread

            self.processed_pages_count = 0 # Reinicia o contador para a nova execu√ß√£o

            for pdf_path in input_files_list:
                if not processing: # Checa a flag global de cancelamento
                    self.log_message("Processamento cancelado antes de concluir todos os PDFs.", "INFO")
                    break

                n_processo = os.path.basename(pdf_path)
                if not n_processo.lower().endswith('.pdf'):
                    self.log_message(f"Pulando arquivo n√£o PDF: {n_processo}", "INFO")
                    continue

                self.log_message(f"Processando documento: {n_processo}", "INFO")
                # Atualiza o status label e barra de progresso para indicar o in√≠cio do arquivo
                self.root.after(0, lambda p=n_processo: self.status_label.config(text=f"Processando: {p}"))
                self.root.update_idletasks() # For√ßa a atualiza√ß√£o da GUI

                ocr_text, num_pages_in_current_pdf = ocr_pdf(pdf_path, temp_dir)

                if not processing: # Checa a flag global de cancelamento novamente ap√≥s OCR
                    self.log_message(f"Processamento de {n_processo} cancelado durante o OCR.", "INFO")
                    break

                self.processed_pages_count += 1 # Conta como 1 PDF processado para a barra
                self.root.after(0, lambda count=self.processed_pages_count, total=self.total_pages_to_process, name=n_processo: self.progress_bar.config(value=count)) # Atualiza a barra por documento
                self.root.after(0, lambda count=self.processed_pages_count, total=self.total_pages_to_process, name=n_processo: self.status_label.config(text=f"Processando documento {count}/{total}: {name}"))
                self.root.update_idletasks() # For√ßa a atualiza√ß√£o da GUI

                total_lines_processed = ws.max_row # Pega o n√∫mero de linhas existentes no Excel

                if ocr_text:
                    info = extract_info(ocr_text)
                    nome_sem_extensao = os.path.splitext(n_processo)[0]

                    if not any(info.values()) and info['cnpj'] == 'N/A' and not info['linhas_digitaveis']:
                        arquivos_com_dados_incompletos.add(nome_sem_extensao)
                        ws.append([nome_sem_extensao, '', '', '', f"Custas{custas_str}:{total_lines_processed:02}"])
                        for cell in ws[ws.max_row]:
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            cell.font = Font(color='000000')
                        error_messages.append(f"Arquivo {n_processo}: Nenhuma informa√ß√£o encontrada.")
                        self.log_message(f"Arquivo {n_processo}: Nenhuma informa√ß√£o encontrada.", "WARNING")
                    elif info['tipo'] == 'guia_custas':
                        if info['cnpj'] != 'N/A' and info['numero_guia'] and info['valor']:
                            ws.append([nome_sem_extensao, info['cnpj'], info['numero_guia'], info['valor'], f"Custas{custas_str}:{total_lines_processed:02}"])
                            self.log_message(f"Arquivo {n_processo}: Guia de Custas processada com sucesso.", "INFO")
                        else:
                            arquivos_com_dados_incompletos.add(nome_sem_extensao)
                            ws.append([nome_sem_extensao, info['cnpj'] if info['cnpj'] else '', info['numero_guia'] if info['numero_guia'] else '', info['valor'] if info['valor'] else '', f"Custas{custas_str}:{total_lines_processed:02}"])
                            for cell in ws[ws.max_row]:
                                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                cell.font = Font(color='000000')
                            error_messages.append(f"Arquivo {n_processo}: Guia de Custas com dados incompletos.")
                            self.log_message(f"Arquivo {n_processo}: Guia de Custas com dados incompletos.", "WARNING")
                    else: # Tipo boleto
                        num_linhas_digitaveis = len(info['linhas_digitaveis'])
                        if num_linhas_digitaveis == 0:
                            arquivos_com_dados_incompletos.add(nome_sem_extensao)
                            ws.append([nome_sem_extensao, info['cnpj'] if info['cnpj'] != 'N/A' else '', '', '', f"Custas{custas_str}:{total_lines_processed:02}"])
                            for cell in ws[ws.max_row]:
                                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                cell.font = Font(color='000000')
                            error_messages.append(f"Arquivo {n_processo}: Nenhuma linha digit√°vel encontrada.")
                            self.log_message(f"Arquivo {n_processo}: Nenhuma linha digit√°vel encontrada.", "WARNING")
                        else:
                            dados_extraidos_com_sucesso = False
                            for i in range(num_linhas_digitaveis):
                                linha_digitavel = info['linhas_digitaveis'][i]
                                valor_monetario = info['valores'][i]
                                if linha_digitavel in linhas_digitaveis_processadas:
                                    self.log_message(f"Linha digit√°vel duplicada encontrada para {n_processo}: {linha_digitavel}", "DEBUG")
                                    continue
                                linhas_digitaveis_processadas.add(linha_digitavel)
                                try:
                                    valor_float = float(valor_monetario.replace(',', '.'))
                                    valor_formatado = "{:,.2f}".format(valor_float).replace(',', '*').replace('.', ',').replace('*', '.')
                                except ValueError:
                                    valor_formatado = valor_monetario
                                    error_messages.append(f"Arquivo {n_processo}: Valor monet√°rio inv√°lido '{valor_monetario}' na linha digit√°vel '{linha_digitavel}'.")
                                    self.log_message(f"Arquivo {n_processo}: Valor monet√°rio inv√°lido '{valor_monetario}'.", "WARNING")

                                obs_nome_arquivo = nome_sem_extensao
                                if i > 0:
                                    obs_nome_arquivo = f"{nome_sem_extensao} - Boleto p√°gina {i + 1}"

                                ws.append([obs_nome_arquivo, info['cnpj'] if info['cnpj'] != 'N/A' else '', linha_digitavel, valor_formatado, f"Custas{custas_str}:{total_lines_processed:02}"])
                                dados_extraidos_com_sucesso = True
                            
                            if not dados_extraidos_com_sucesso:
                                arquivos_com_dados_incompletos.add(nome_sem_extensao)
                                ws.append([nome_sem_extensao, info['cnpj'] if info['cnpj'] != 'N/A' else '', '', '', f"Custas{custas_str}:{total_lines_processed:02}"])
                                for cell in ws[ws.max_row]:
                                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                    cell.font = Font(color='000000')
                                error_messages.append(f"Arquivo {n_processo}: Nenhuma linha digit√°vel v√°lida/nova encontrada.")
                                self.log_message(f"Arquivo {n_processo}: Nenhuma linha digit√°vel v√°lida/nova encontrada.", "WARNING")

                        if info['cnpj'] == 'N/A' and num_linhas_digitaveis > 0:
                            arquivos_com_dados_incompletos.add(nome_sem_extensao)
                            for cell in ws[ws.max_row]:
                                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                    cell.font = Font(color='000000')
                            error_messages.append(f"Arquivo {n_processo}: CNPJ n√£o encontrado ou ignorado.")
                            self.log_message(f"Arquivo {n_processo}: CNPJ n√£o encontrado ou ignorado.", "WARNING")

                    for col in range(1, ws.max_column + 1):
                        column_letter = openpyxl.utils.get_column_letter(col)
                        column_width = max(len(str(cell.value)) if cell.value else 0 for cell in ws[column_letter]) + 2
                        ws.column_dimensions[column_letter].width = max(column_width, 10)

                    if num_pages_in_current_pdf > 1:
                        arquivos_com_paginas_a_mais.add(nome_sem_extensao)
                        self.log_message(f"Arquivo {n_processo}: Possui mais de uma p√°gina.", "WARNING")
                else:
                    error_messages.append(f"Arquivo {n_processo}: Falha no processamento do OCR.")
                    arquivos_com_dados_incompletos.add(nome_sem_extensao)
                    ws.append([nome_sem_extensao, '', '', '', f"Custas{custas_str}:{total_lines_processed:02}"])
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        cell.font = Font(color='000000')
                    self.log_message(f"Arquivo {n_processo}: Falha no processamento do OCR.", "ERROR")

            if processing: # Se n√£o foi cancelado durante o loop principal
                for row_idx in range(2, ws.max_row + 1):
                    nome_arquivo_celula = ws.cell(row_idx, 1).value
                    if isinstance(nome_arquivo_celula, str):
                        nome_base_arquivo = nome_arquivo_celula.split(" - Boleto p√°gina ")[0]
                        if nome_base_arquivo in arquivos_com_dados_incompletos:
                            for cell in ws[row_idx]:
                                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                cell.font = Font(color='000000')

                for cell in ws['D']:
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                ws['D1'].alignment = Alignment(horizontal='left')

                for row_idx in range(2, ws.max_row + 1):
                    valor_boleto_cell = ws.cell(row_idx, 4)
                    valor_boleto = valor_boleto_cell.value
                    if valor_boleto and isinstance(valor_boleto, str):
                        try:
                            valor_boleto_float = float(valor_boleto.replace('.', '').replace(',', '.'))
                            if valor_boleto_float > 2000:
                                error_messages.append(f"Arquivo {ws.cell(row_idx, 1).value}: Valor do boleto (R$ {valor_boleto}) acima de R$ 2000. Verificar manual.")
                                self.log_message(f"Arquivo {ws.cell(row_idx, 1).value}: Valor do boleto (R$ {valor_boleto}) acima de R$ 2000. Verificar manual.", "WARNING")
                                for cell in ws[row_idx]:
                                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                    cell.font = Font(color='000000')
                        except ValueError:
                            pass # J√° tratado se o valor n√£o √© num√©rico
                
                # Salva o arquivo final
                wb.save(result_file_str)
                self.log_message(f"Arquivo Excel salvo em: {result_file_str}", "INFO")

                num_erros_reportados = len(error_messages) + len(arquivos_com_paginas_a_mais) + len(arquivos_com_dados_incompletos)
                num_registros_extraidos = ws.max_row - 1

                if save_csv_bool and num_erros_reportados == 0 and num_registros_extraidos > 0:
                    self.save_to_csv_method(result_file_str, ws) # Chama o m√©todo da classe
                    self.log_message(f"Arquivo CSV salvo em: {os.path.splitext(result_file_str)[0] + '.csv'}", "INFO")
                elif save_csv_bool:
                    self.log_message("CSV autom√°tico n√£o criado devido a diverg√™ncias ou falta de dados.", "WARNING")

        except Exception as e:
            logger.exception("Erro cr√≠tico durante o processamento dos PDFs na thread")
            self.log_message(f"Erro cr√≠tico durante o processamento: {e}", "CRITICAL_ERROR")
            error_messages.append(f"Erro cr√≠tico: {e}")
        finally:
            temp_dir_obj.cleanup()
            self.log_message("Pasta tempor√°ria de OCR removida.", "DEBUG")

            # Garante que a barra chegue a 100% mesmo se a contagem inicial for imprecisa
            # Se o processamento n√£o foi cancelado, ela deve atingir o valor m√°ximo do documento.
            if processing: # Se a flag global ainda √© True, significa que terminou com sucesso (n√£o foi cancelado)
                self.root.after(0, lambda: self.progress_bar.config(value=self.total_pages_to_process, maximum=self.total_pages_to_process))
                self.root.after(100, lambda: self.status_label.config(text=f"Processamento conclu√≠do! Gerando relat√≥rio..."))
            
            # Chama a fun√ß√£o de finaliza√ß√£o na thread principal
            self.root.after(150, lambda: self._processing_complete(
                error_messages, list(arquivos_com_paginas_a_mais), list(arquivos_com_dados_incompletos)
            ))


    def _processing_complete(self, error_messages, arquivos_com_paginas_a_mais_list, arquivos_com_dados_incompletos_list):
        """Finaliza o processamento, atualiza a GUI e mostra popups."""
        global processing # Acessa a flag global

        # Se o processamento foi cancelado explicitamente (flag global `processing` √© False NO IN√çCIO DESTA FUN√á√ÉO)
        # O `processing` √© definido como False em `cancel_processing_gui` ou em `_actual_processing_task` em caso de erro cr√≠tico.
        if not processing: # Esta condi√ß√£o deve ser a primeira para tratar o cancelamento.
            self.log_message("Processamento cancelado pelo usu√°rio ou finalizado por erro.", "INFO")
            self.show_cancelled_popup_gui()
            self.status_label.config(text="Processamento cancelado.")
            self.set_progress_bar_style("Error.Horizontal.TProgressbar")
            self._update_main_button_state('initial') # Volta o bot√£o para "Iniciar" e reativa
            return

        # Se n√£o foi cancelado, verifica o resultado
        num_erros_reportados = len(error_messages) + len(arquivos_com_paginas_a_mais_list) + len(arquivos_com_dados_incompletos_list)
        
        num_registros_extraidos = 0
        if os.path.exists(self.result_file): # Usa self.result_file
            try:
                wb_final = openpyxl.load_workbook(self.result_file)
                ws_final = wb_final.active
                num_registros_extraidos = ws_final.max_row - 1
            except Exception as e:
                self.log_message(f"Erro ao contar registros no arquivo final: {e}", "WARNING")

        if num_erros_reportados == 0 and num_registros_extraidos > 0:
            self.set_progress_bar_style("Success.Horizontal.TProgressbar")
            self.status_label.config(text="Conclu√≠do com sucesso!")
            self.log_message("Processamento conclu√≠do com sucesso!", "SUCCESS")
            self.show_success_popup_gui(self.result_file, self.save_csv_var.get())
        elif num_registros_extraidos == 0 and num_erros_reportados == 0:
            self.set_progress_bar_style("Error.Horizontal.TProgressbar")
            self.status_label.config(text="Conclu√≠do (Sem dados extra√≠dos).")
            self.log_message("Processamento conclu√≠do, mas nenhum dado foi extra√≠do.", "WARNING")
            self.show_divergencia_popup_gui(error_messages, self.result_file, arquivos_com_paginas_a_mais_list, arquivos_com_dados_incompletos_list, self.save_csv_var.get(), no_data=True)
        elif num_erros_reportados > 0:
            self.set_progress_bar_style("Error.Horizontal.TProgressbar")
            self.status_label.config(text="Conclu√≠do com ERROS/DIVERG√äNCIAS.")
            self.log_message(f"Processamento conclu√≠do com {num_erros_reportados} problemas/erros.", "WARNING")
            self.show_divergencia_popup_gui(error_messages, self.result_file, arquivos_com_paginas_a_mais_list, arquivos_com_dados_incompletos_list, self.save_csv_var.get(), no_data=False)
        else:
            self.set_progress_bar_style("Default.Horizontal.TProgressbar")
            self.status_label.config(text="Processamento finalizado.")
            self.log_message("Processamento finalizado em estado indefinido.", "INFO")
        
        # Sempre retorna o bot√£o para o estado inicial ap√≥s a conclus√£o (seja sucesso, erro ou sem dados)
        self._update_main_button_state('initial')


    def select_input_files(self):
        self.input_files = filedialog.askopenfilenames(
            filetypes=[("Arquivos PDF", "*.pdf")],
            title="Selecione os arquivos PDF"
        )
        if self.input_files:
            self.input_files = list(self.input_files)
            self.input_dir = os.path.dirname(self.input_files[0])
            self.input_dir_label.config(text=f"{len(self.input_files)} arquivos selecionados")
            self.log_message(f"Selecionados {len(self.input_files)} arquivos PDF.", "INFO")
        else:
            self.input_files = []
            self.input_dir = ""
            self.input_dir_label.config(text="PDF n√£o selecionado")
            self.log_message("Sele√ß√£o de PDFs cancelada.", "INFO")

    def select_result_file(self):
        self.result_file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Planilhas Excel", "*.xlsx")],
            title="Salve a planilha de resultados",
        )
        if self.result_file:
            self.result_file_label.config(text=f"{os.path.basename(self.result_file)}")
            self.log_message(f"Arquivo de resultados selecionado: {os.path.basename(self.result_file)}", "INFO")
        else:
            self.result_file = ""
            self.result_file_label.config(text="Arquivo n√£o selecionado")
            self.log_message("Sele√ß√£o do arquivo de resultados cancelada.", "INFO")

    def start_processing(self):
        # Esta verifica√ß√£o agora √© mais gen√©rica, pois o bot√£o pode estar desabilitado
        # temporariamente (ex: se clicou "Cancelar" mas a thread n√£o terminou)
        if self.process_button['state'] == DISABLED:
            self.log_message("Bot√£o de processamento desabilitado. Aguarde ou reinicie.", "WARNING")
            return

        self.log_message("Iniciando o processamento...", "INFO")

        try:
            if not self.input_files:
                self.log_message("Arquivos PDF n√£o selecionados.", "ERROR")
                messagebox.showerror("Erro", "Arquivos n√£o Selecionados: Selecione os arquivos PDF.", icon="error")
                return

            if not self.result_file:
                self.log_message("Planilha de resultado n√£o selecionada.", "ERROR")
                messagebox.showerror("Erro", "Planilha n√£o Selecionada: Selecione a planilha de resultados.", icon="error")
                return

            if not os.path.exists(poppler_path):
                self.log_message(f"Pasta do Poppler n√£o encontrada: {poppler_path}", "CRITICAL_ERROR")
                messagebox.showerror("Erro", f"Pasta do Poppler n√£o encontrada em: {poppler_path}. O programa n√£o poder√° funcionar corretamente", icon="error")
                return

            custas = self.custas_entry.get()
            if not re.match(r'^[0-9\.\:\/\\]{0,5}$', custas):
                self.log_message(f"Valor de custas inv√°lido: {custas}", "WARNING")
                messagebox.showerror("Erro", "Digite um valor v√°lido para as custas (apenas n√∫meros, '.', ':', '/', '\\) com at√© 5 caracteres.", icon="error")
                return

            # Atualiza o estado do bot√£o para "Cancelar" e habilita
            self._update_main_button_state('processing_start')
            self.log_message("Configura√ß√µes validadas. Preparando para processar.", "INFO")

            output_dir = self.input_dir if self.input_dir else os.getcwd()
            self.log_message(f"Pasta de sa√≠da para arquivos tempor√°rios: {output_dir}", "DEBUG")

            thread = Thread(target=self._actual_processing_task, args=(self.input_files, output_dir, self.result_file, custas, self.save_csv_var.get()))
            thread.start()

        except Exception as e:
            self.log_message(f"Erro inesperado antes de iniciar a thread: {e}", "CRITICAL_ERROR")
            messagebox.showerror("Erro", f"Erro inesperado: {e}", icon="error")
            self._update_main_button_state('initial') # Retorna para o estado inicial em caso de erro na inicializa√ß√£o

    def cancel_processing_gui(self):
        """Fun√ß√£o para cancelar o processamento, chamada pelo bot√£o principal quando est√° em modo 'cancelar'."""
        global processing # Acessa a vari√°vel global de controle de processamento
        if processing: # S√≥ envia o sinal se o processamento estiver ativo
            processing = False # Define a flag global para parar a thread
            self._update_main_button_state('cancelling_pending') # Atualiza o bot√£o para "Cancelando..." e o desabilita

    def show_divergencia_popup_gui(self, error_messages, result_file, arquivos_com_paginas_a_mais, arquivos_com_dados_incompletos, save_csv, no_data=False):
        popup_div = Toplevel(self.root)
        popup_div.title("Processamento Conclu√≠do com Diverg√™ncias")
        popup_div.transient(self.root)
        popup_div.grab_set()
        popup_div.resizable(False, False)

        if no_data:
            label = Label(popup_div, text="Processamento conclu√≠do! Nenhum dado extra√≠do.", font=("Segoe UI Bold", 10), fg="red")
        else:
            label = Label(popup_div, text="Processamento conclu√≠do! Diverg√™ncias encontradas:", font=("Segoe UI Bold", 10), fg="#FF0000")
        label.pack(pady=10)

        if arquivos_com_paginas_a_mais:
            label_paginas_a_mais = Label(popup_div, text=f"Arquivos com mais de uma p√°gina: {len(arquivos_com_paginas_a_mais)}", font=("Segoe UI Bold", 10))
            label_paginas_a_mais.pack(pady=5)
            frame_paginas = Frame(popup_div)
            scrollbar_paginas = Scrollbar(frame_paginas, orient=VERTICAL)
            lista_paginas_text = Text(frame_paginas, wrap=WORD, height=min(5, len(arquivos_com_paginas_a_mais)), width=50, yscrollcommand=scrollbar_paginas.set, font=("Segoe UI", 9))
            for item in arquivos_com_paginas_a_mais:
                lista_paginas_text.insert(END, item + "\n")
            lista_paginas_text.config(state=DISABLED)
            scrollbar_paginas.config(command=lista_paginas_text.yview)
            scrollbar_paginas.pack(side=RIGHT, fill=Y)
            lista_paginas_text.pack(side=LEFT, fill=BOTH, expand=True)
            frame_paginas.pack(pady=5, padx=10, fill=X)

        if arquivos_com_dados_incompletos:
            label_dados_incompletos = Label(popup_div, text=f"Arquivos com informa√ß√µes faltando: {len(arquivos_com_dados_incompletos)}", font=("Segoe UI Bold", 10))
            label_dados_incompletos.pack(pady=5)
            frame_dados = Frame(popup_div)
            scrollbar_dados = Scrollbar(frame_dados, orient=VERTICAL)
            lista_dados_text = Text(frame_dados, wrap=WORD, height=min(5, len(arquivos_com_dados_incompletos)), width=50, yscrollcommand=scrollbar_dados.set, font=("Segoe UI", 9))
            for item in arquivos_com_dados_incompletos:
                lista_dados_text.insert(END, item + "\n")
            lista_dados_text.config(state=DISABLED)
            scrollbar_dados.config(command=lista_dados_text.yview)
            scrollbar_dados.pack(side=RIGHT, fill=Y)
            lista_dados_text.pack(side=LEFT, fill=BOTH, expand=True)
            frame_dados.pack(pady=5, padx=10, fill=X)

        if error_messages:
            label_erros_gerais = Label(popup_div, text="Outras observa√ß√µes:", font=("Segoe UI Bold", 10))
            label_erros_gerais.pack(pady=5)
            frame_erros = Frame(popup_div)
            scrollbar_erros = Scrollbar(frame_erros, orient=VERTICAL)
            lista_erros_text = Text(frame_erros, wrap=WORD, height=min(5, len(error_messages)), width=50, yscrollcommand=scrollbar_erros.set, font=("Segoe UI", 9))
            for msg in error_messages:
                lista_erros_text.insert(END, msg + "\n")
            lista_erros_text.config(state=DISABLED)
            scrollbar_erros.config(command=lista_erros_text.yview)
            scrollbar_erros.pack(side=RIGHT, fill=Y)
            lista_erros_text.pack(side=LEFT, fill=BOTH, expand=True)
            frame_erros.pack(pady=5, padx=10, fill=X)

        if save_csv and (error_messages or arquivos_com_paginas_a_mais or arquivos_com_dados_incompletos):
            label_csv_nao_criado = Label(popup_div, text="(CSV autom√°tico n√£o criado, por falta de confiabilidade)", font=("Segoe UI Bold", 10), fg="#FF0000")
            label_csv_nao_criado.pack(pady=5)

        button_frame = Frame(popup_div)
        button_frame.pack(pady=10)

        ok_button = self.create_button(button_frame, "OK", lambda: [popup_div.destroy(), webbrowser.open_new_tab(f"file://{result_file}")], is_default=True, fg_color="#000000")
        ok_button.pack(side=LEFT, padx=10)

        todos_arquivos_para_conferir = list(set(arquivos_com_paginas_a_mais + arquivos_com_dados_incompletos))

        if todos_arquivos_para_conferir:
            conferir_pdfs_button = self.create_button(button_frame, "Conferir PDFs",
                                      command=lambda: [popup_div.destroy(), self.abrir_arquivos_pdf(todos_arquivos_para_conferir), webbrowser.open_new_tab(f"file://{result_file}")],
                                      fg_color="#0000FF")
            conferir_pdfs_button.pack(side=LEFT, padx=10)

        center_window(popup_div)

    def show_success_popup_gui(self, result_file, save_csv):
        popup_success = Toplevel(self.root)
        popup_success.title("Processamento Conclu√≠do com Sucesso")
        popup_success.transient(self.root)
        popup_success.grab_set()
        popup_success.resizable(False, False)

        label = Label(popup_success, text="Processamento conclu√≠do com sucesso!", font=("Segoe UI Bold", 10), fg="green")
        label.pack(pady=10)

        if save_csv:
            label_csv_criado = Label(popup_success, text="(CSV autom√°tico tamb√©m foi criado)", font=("Segoe UI Bold", 10), fg="green")
            label_csv_criado.pack(pady=5)

        button_frame = Frame(popup_success)
        button_frame.pack(pady=10)

        ok_button = self.create_button(button_frame, "OK", lambda: [popup_success.destroy(), webbrowser.open_new_tab(f"file://{result_file}")], is_default=True, fg_color="#000000")
        ok_button.pack(side=LEFT, padx=10)

        center_window(popup_success)

    def show_cancelled_popup_gui(self):
        popup_cancel = Toplevel(self.root)
        popup_cancel.title("Processamento Cancelado!")
        popup_cancel.transient(self.root)
        popup_cancel.grab_set()
        popup_cancel.resizable(False, False)

        label = Label(popup_cancel, text="Processamento Cancelado!", font=("Segoe UI Bold", 10), fg="red")
        label.pack(pady=10)

        button_frame = Frame(popup_cancel)
        button_frame.pack(pady=10)

        ok_button = self.create_button(button_frame, "OK", popup_cancel.destroy, is_default=True, fg_color="#FF0000")
        ok_button.pack(side=LEFT, padx=10)

        center_window(popup_cancel)

    def show_info(self):
        """Exibe uma caixa de di√°logo com informa√ß√µes e op√ß√µes de debug/configura√ß√£o."""
        info_popup = Toplevel(self.root)
        info_popup.title("Informa√ß√£o")
        info_popup.transient(self.root)
        info_popup.grab_set()
        info_popup.resizable(False, False)
        info_popup.configure(bg="#f0f0f0") # Cor de fundo padr√£o

        # Frame principal para conte√∫do
        content_frame = Frame(info_popup, padx=15, pady=15, bg=info_popup.cget("bg"))
        content_frame.pack(expand=True, fill=BOTH)

        version_label = Label(content_frame, text=f"{self.root.title()} - by Elias", font=("Segoe UI", 10), bg=content_frame.cget("bg"), fg="#002b00")
        version_label.pack(pady=(0,5))
        pix_label = Label(content_frame, text="Chamado via mensagem Pix: eliasgkersten@gmail.com", font=("Segoe UI", 10), bg=content_frame.cget("bg"), fg="#002b00")
        pix_label.pack(pady=5)

        button_frame = Frame(content_frame, bg=content_frame.cget("bg"))
        button_frame.pack(pady=(10,0))

        # Bot√£o Debug (verde)
        debug_button = Button(button_frame, text="Debug", command=self.open_log_file, font=("Segoe UI Bold", 10), bg="#4CAF50", fg="white", relief=FLAT, padx=10, pady=5)
        debug_button.pack(side=LEFT, padx=5)

        # Bot√£o Excluir Debug (vermelho)
        c_debug_button = Button(button_frame, text="üóëÔ∏èExcluir Debug", command=self.delete_log_file, font=("Segoe UI Bold", 10), bg="#F32121", fg="white", relief=FLAT, padx=10, pady=5)
        c_debug_button.pack(side=LEFT, padx=5)

        # Bot√£o Filtro CNPJ (cinza azulado)
        config_button = Button(button_frame, text="üõ†Ô∏è Filtro", command=lambda: self.show_filtro_cnpj_config_popup(info_popup), font=("Segoe UI Bold", 10), bg="#607D8B", fg="white", relief=FLAT, padx=10, pady=5)
        config_button.pack(side=LEFT, padx=5)

        # Bot√£o Sair (preto)
        exit_button = Button(button_frame, text="Sair", command=info_popup.destroy, font=("Segoe UI Bold", 10), bg="#000000", fg="white", relief=FLAT, padx=10, pady=5)
        exit_button.pack(side=LEFT, padx=5)

        center_window(info_popup)
        self.log_message("Informa√ß√µes do programa exibidas.", "INFO")

    def abrir_arquivos_pdf(self, arquivos_nomes_base):
        # input_files agora √© self.input_files
        arquivos_abertos = 0
        if not self.input_files:
            self.log_message("Tentativa de abrir PDFs sem input_files definidos.", "WARNING")
            return

        for nome_base in arquivos_nomes_base:
            caminho_completo = next((f for f in self.input_files if os.path.splitext(os.path.basename(f))[0] == nome_base), None)
            if caminho_completo:
                try:
                    webbrowser.open_new_tab(f"file://{caminho_completo}")
                    arquivos_abertos +=1
                except Exception as e:
                    self.log_message(f"Erro ao tentar abrir PDF {caminho_completo}: {e}", "ERROR")
            else:
                self.log_message(f"Arquivo PDF com nome base '{nome_base}' n√£o encontrado na lista de input_files.", "WARNING")
        if arquivos_abertos == 0 and arquivos_nomes_base:
            messagebox.showwarning("Aviso", "N√£o foi poss√≠vel localizar os arquivos PDF para abrir. Verifique se foram movidos ou renomeados.", icon="warning")

    def create_button(self, parent, text, command, is_default=False, fg_color="#FF0000"):
        button = ttk.Button(parent, text=text, command=command)
        button_style_name = f"{text.replace(' ', '_').replace('\'','')}.TButton"
        style = ttk.Style()
        # Se o estilo j√° existe, configure-o. Se n√£o, crie-o.
        # Evita erro se o nome do estilo for din√¢mico e j√° tiver sido criado antes.
        try:
            style.configure(button_style_name, font=("Segoe UI Bold", 10), background=fg_color, foreground="white")
            style.map(button_style_name, background=[('active', fg_color)])
        except TclError: # Pode ocorrer se o estilo j√° foi configurado de forma diferente
            pass # Permite que o estilo padr√£o seja usado se o nome j√° estiver em uso.
        button.config(style=button_style_name)

        if is_default:
            button.bind("<Return>", lambda event: command())
            button.focus_set()
        return button

    def open_log_file(self):
        try:
            webbrowser.open_new_tab(f"file://{log_file_path}")
            self.log_message("Arquivo de log aberto.", "INFO")
        except Exception as e:
            messagebox.showerror("Erro", f"N√£o foi poss√≠vel abrir o arquivo de log: {e}", icon="error")
            self.log_message(f"Erro ao abrir arquivo de log: {e}", "ERROR")

    def delete_log_file(self):
        try:
            # Primeiro, remova o handler de arquivo existente
            for handler in logger.handlers[:]:
                if isinstance(handler, logging.FileHandler):
                    logger.removeHandler(handler)
                    handler.close()
            
            logging.shutdown()

            if os.path.exists(log_file_path):
                os.remove(log_file_path)
                self.log_message("Arquivo de log deletado pelo usu√°rio.", "INFO")
                messagebox.showinfo("Sucesso", "Arquivo de log deletado com sucesso.", icon="info")
                self.log_text.config(state=tk.NORMAL)
                self.log_text.delete(1.0, tk.END)
                self.log_text.config(state=tk.DISABLED)
            else:
                messagebox.showerror("Erro", "Arquivo de log n√£o encontrado para deletar.", icon="error")
                self.log_message("Arquivo de log n√£o encontrado para deletar (ap√≥s tentativa de remo√ß√£o).", "ERROR")
                
            # Recria o handler de arquivo AP√ìS deletar e confirmar
            file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
            file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
            logger.addHandler(file_handler)

        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo de log n√£o encontrado para deletar.", icon="error")
            self.log_message("Arquivo de log n√£o encontrado para deletar (ap√≥s tentativa de remo√ß√£o).", "ERROR")
        except PermissionError:
            messagebox.showerror("Erro", "N√£o foi poss√≠vel deletar o arquivo de log. Ele pode estar em uso ou voc√™ n√£o tem permiss√£o.", icon="error")
            self.log_message("Erro de permiss√£o ao deletar arquivo de log.", "ERROR")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao deletar o arquivo de log: {e}", icon="error")
            self.log_message(f"Erro inesperado ao deletar arquivo de log: {e}", "ERROR")

    def show_filtro_cnpj_config_popup(self, parent_window):
        global ignored_cnpjs_list

        config_popup = Toplevel(parent_window)
        config_popup.title("Configurar Filtro de CNPJ")
        config_popup.transient(parent_window)
        config_popup.grab_set()
        config_popup.resizable(False, False)
        config_popup.configure(bg="#f0f0f0")

        main_frame = Frame(config_popup, padx=15, pady=15, bg=config_popup.cget("bg"))
        main_frame.pack(fill=BOTH, expand=True)

        Label(main_frame, text="CNPJs a serem ignorados (separados por v√≠rgula):", font=("Segoe UI", 10), bg=main_frame.cget("bg")).pack(pady=(0,5), anchor="w")

        cnpj_entry_var = StringVar()
        cnpj_entry_var.set(",".join(ignored_cnpjs_list))

        cnpj_entry = Entry(main_frame, textvariable=cnpj_entry_var, width=60, font=("Segoe UI", 10))
        cnpj_entry.pack(pady=5, fill=X)

        button_frame = Frame(main_frame, bg=main_frame.cget("bg"))
        button_frame.pack(pady=(10,0), fill=X)

        def on_save():
            cnpjs_text = cnpj_entry_var.get()
            cnpjs_to_save = []
            has_invalid = False
            for c in cnpjs_text.split(','):
                c = c.strip()
                if c:
                    if re.match(r'^\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}$', c):
                        cnpjs_to_save.append(c)
                    else:
                        has_invalid = True
            
            if has_invalid:
                if not messagebox.askyesno("CNPJ Inv√°lido", "Um ou mais CNPJs parecem estar em formato inv√°lido. Deseja salvar mesmo assim?", icon="warning", parent=config_popup):
                    return

            final_cnpjs_string = ",".join(cnpjs_to_save)
            if save_cnpjs_to_config(final_cnpjs_string):
                messagebox.showinfo("Sucesso", "Filtro de CNPJs salvo com sucesso!", parent=config_popup)
                self.log_message("Filtro de CNPJs salvo via GUI.", "INFO")
                config_popup.destroy()

        def on_open_config_file():
            try:
                if not os.path.exists(filtro_config_path):
                    save_cnpjs_to_config(cnpj_entry_var.get())
                
                if sys.platform == "win32":
                    os.startfile(filtro_config_path)
                elif sys.platform == "darwin":
                    subprocess.call(["open", filtro_config_path])
                else:
                    subprocess.call(["xdg-open", filtro_config_path])
                messagebox.showinfo("Informa√ß√£o", f"Ap√≥s editar e salvar o arquivo '{os.path.basename(filtro_config_path)}'(Caso tenha feito fora do Editor do programa), clique em 'cancelar' nesta janela para aplicar as mudan√ßas", parent=config_popup, icon="info")
                self.log_message("Arquivo de configura√ß√£o de CNPJ aberto externamente.", "INFO")
            except Exception as e:
                logger.error(f"Erro ao abrir arquivo de configura√ß√£o de CNPJ: {e}")
                messagebox.showerror("Erro", f"N√£o foi poss√≠vel abrir o arquivo de configura√ß√£o: {e}", icon="error", parent=config_popup)
                self.log_message(f"Erro ao abrir arquivo de configura√ß√£o de CNPJ: {e}", "ERROR")

        save_button = self.create_button(button_frame, "Salvar", on_save, fg_color="#4CAF50")
        save_button.pack(side=LEFT, padx=5)

        open_config_button = self.create_button(button_frame, "Abrir Filtro.config", on_open_config_file, fg_color="#2196F3")
        open_config_button.pack(side=LEFT, padx=5)

        cancel_button = self.create_button(button_frame, "Cancelar", config_popup.destroy, fg_color="#F44336")
        cancel_button.pack(side=RIGHT, padx=5)

        center_window(config_popup)
        config_popup.focus_set()
        cnpj_entry.focus()

    def toggle_csv_save(self):
        self.save_csv_var.set(not self.save_csv_var.get())
        self.log_message(f"Op√ß√£o 'CSV ponto e v√≠rgula' {'ativada' if self.save_csv_var.get() else 'desativada'}.", "INFO")

    def save_to_csv_method(self, result_file, ws):
        csv_file = os.path.splitext(result_file)[0] + ".csv"
        with open(csv_file, 'w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file, delimiter=';')
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

class TextLogHandler(logging.Handler):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        self.queue = []
        self.text_widget.after(100, self.periodic_check)

    def emit(self, record):
        msg = self.format(record)
        level_name = record.levelname.upper()
        self.queue.append((msg, level_name))

    def periodic_check(self):
        while self.queue:
            msg, level_name = self.queue.pop(0)
            self.text_widget.config(state=tk.NORMAL)
            self.text_widget.insert(tk.END, f"{msg}\n", level_name)
            self.text_widget.config(state=tk.DISABLED)
            self.text_widget.see(tk.END)
        self.text_widget.after(100, self.periodic_check)

if __name__ == "__main__":
    root = tk.Tk()
    app = AppCelescReporter(root) # Cria a inst√¢ncia da aplica√ß√£o

    # Vari√°vel global `processing` para ser usada por fun√ß√µes que n√£o s√£o m√©todos da classe
    # como `ocr_pdf` para permitir o cancelamento.
    processing = False # Inicializado aqui e controlado por `_update_main_button_state`

    # Garante que o bot√£o principal esteja no estado "Iniciar" ao iniciar a aplica√ß√£o
    app._update_main_button_state('initial')

    if getattr(sys, 'frozen', False):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW

    root.mainloop()