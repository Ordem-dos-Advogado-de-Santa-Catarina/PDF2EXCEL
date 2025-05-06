import os
import subprocess
import sys
import re
from tkinter import *
from tkinter import filedialog, messagebox, ttk, font
from pdf2image import convert_from_path
import pytesseract
import openpyxl
from openpyxl import Workbook
import webbrowser
from threading import Thread
import logging
import tkinter as tk
import glob  # Importe o módulo glob para listar arquivos PDF
from openpyxl.styles import Alignment, PatternFill, Font
import csv # Importe o módulo csv
import tempfile  # Importe o módulo tempfile para pastas temporárias
import shutil # Importe o módulo shutil para remover pastas

# *** Configuração de Logging ***
log_dir = os.path.join(os.environ['APPDATA'], 'PDF2EXCEL')
if not os.path.exists(log_dir):
    os.makedirs(log_dir)
log_file_path = os.path.join(log_dir, 'PDF2EXCEL.log')

logging.basicConfig(filename=log_file_path,
                    level=logging.INFO, # ou logging.DEBUG para mais detalhes
                    filemode='w', # 'w' para sobrescrever o arquivo a cada execução
                    format='%(asctime)s - %(levelname)s - %(message)s')

logging.info("Programa PDF2EXCEL iniciado.")

# *** Configuração do Filtro de CNPJ ***
filtro_config_path = os.path.join(log_dir, 'Filtro.config')
ignored_cnpjs_list = []
DEFAULT_IGNORED_CNPJ = "82.519.190/0001-12" # CNPJ da OAB como padrão

def load_ignored_cnpjs():
    global ignored_cnpjs_list
    try:
        if os.path.exists(filtro_config_path):
            with open(filtro_config_path, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if content:
                    ignored_cnpjs_list = [cnpj.strip() for cnpj in content.split(',') if cnpj.strip()]
                else:
                    # Se o arquivo existe mas está vazio, usa o padrão e salva
                    ignored_cnpjs_list = [DEFAULT_IGNORED_CNPJ]
                    save_cnpjs_to_config(DEFAULT_IGNORED_CNPJ) # Salva o padrão se o arquivo estiver vazio
        else:
            # Se o arquivo não existe, cria com o padrão
            ignored_cnpjs_list = [DEFAULT_IGNORED_CNPJ]
            save_cnpjs_to_config(DEFAULT_IGNORED_CNPJ) # Salva o padrão se o arquivo não existir
    except Exception as e:
        logging.error(f"Erro ao carregar CNPJs ignorados: {e}")
        ignored_cnpjs_list = [DEFAULT_IGNORED_CNPJ] # Fallback para o padrão em caso de erro
    logging.info(f"CNPJs ignorados carregados: {ignored_cnpjs_list}")

def save_cnpjs_to_config(cnpjs_string):
    global ignored_cnpjs_list
    try:
        with open(filtro_config_path, 'w', encoding='utf-8') as f:
            f.write(cnpjs_string)
        # Recarrega a lista após salvar
        load_ignored_cnpjs()
        logging.info(f"CNPJs ignorados salvos em {filtro_config_path}: {cnpjs_string}")
        return True
    except Exception as e:
        logging.error(f"Erro ao salvar CNPJs no arquivo de configuração: {e}")
        messagebox.showerror("Erro", f"Não foi possível salvar o arquivo de configuração de CNPJs: {e}", icon="error")
        return False

# Carrega os CNPJs ignorados na inicialização
load_ignored_cnpjs()


# Definir o caminho do Poppler (apenas para a versão .exe)
poppler_path = os.path.join(sys._MEIPASS, 'poppler', 'bin') if getattr(sys, 'frozen', False) else r"C:\Program Files\poppler\bin"

# Se o Poppler não estiver na pasta padrão do executável, verifica se está instalado em C:\Program Files\poppler\bin
if getattr(sys, 'frozen', False) and not os.path.exists(poppler_path):
    poppler_path = r"C:\Program Files\poppler\bin"  # Define o caminho alternativo

# Definir o caminho do Tesseract OCR (apenas para a versão .exe)
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Centraliza o programa na tela
def center_window(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    window.geometry(f'+{x}+{y}')

# Define no path do windows o poppler
def ocr_pdf(pdf_path, temp_dir): # Adicionado temp_dir
    try:
        images = convert_from_path(pdf_path, poppler_path=poppler_path, output_folder=temp_dir, paths_only=False, fmt='jpeg')
        num_pages = len(images)
        text = ""
        for image in images:
            if not processing:
                break
            text += pytesseract.image_to_string(image, lang='por')
        return text, num_pages
    except Exception as e:
        logging.exception(f"Erro ao processar OCR do PDF: {pdf_path}")
        return None, 0

def extract_info(text):
    global ignored_cnpjs_list # Usa a lista global de CNPJs ignorados
    cnpj = None
    linhas_digitaveis = []
    valores_monetarios = []
    numero_guia = None
    valor = None

    # Busca o CNPJ do Beneficiário (modificado para usar a lista de ignorados)
    cnpj_matches = re.findall(r'(\d{2}\.\d{3}\.\d{3}\/\d{4}\-\d{2})', text)
    valid_cnpjs = [cnp for cnp in cnpj_matches if cnp not in ignored_cnpjs_list] # Filtra usando a lista
    if valid_cnpjs:
        cnpj = valid_cnpjs[0]
    else:
        cnpj = 'N/A'

    if "GUIA ÚNICA DE CUSTAS" in text:
        numero_guia_match = re.search(r"Nº da Guia\s*([\d\.]+/\d+)", text)
        if numero_guia_match:
            numero_guia = numero_guia_match.group(1)
        valor_match = re.search(r"R\$\s*([\d,.]+)", text)
        if valor_match:
            valor = valor_match.group(1)
        return {
            'cnpj': cnpj,
            'numero_guia': numero_guia,
            'valor': valor,
            'linhas_digitaveis': [],
            'valores': [],
            'tipo': 'guia_custas'
        }
    else:
        text = re.sub(r'\d{3}-\d', '', text)
        for line in text.splitlines():
            cleaned_line = re.sub(r'[^0-9]', '', line)
            if 47 <= len(cleaned_line) <= 48:
                linhas_digitaveis.append(cleaned_line)
                valor_monetario = f"{cleaned_line[-10:-2]},{cleaned_line[-2:]}"
                valores_monetarios.append(valor_monetario)
        return {
            'cnpj': cnpj,
            'linhas_digitaveis': linhas_digitaveis,
            'valores': valores_monetarios,
            'numero_guia': None,
            'valor': None,
            'tipo': 'boleto'
        }

arquivos_com_paginas_a_mais = set()
arquivos_com_dados_incompletos = set()

def save_to_csv(result_file, ws):
    csv_file = os.path.splitext(result_file)[0] + ".csv"
    with open(csv_file, 'w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file, delimiter=';')
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

def process_pdfs(input_files, output_dir, result_file, custas, save_csv):
    global arquivos_com_paginas_a_mais, arquivos_com_dados_incompletos, processing

    arquivos_com_paginas_a_mais.clear()
    arquivos_com_dados_incompletos.clear()

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if not os.path.exists(result_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Boletos"
        ws.append(['Obeservação', 'Fornecedor', 'Código de Barras', 'Valor', 'Nome do Titulo'])
    else:
        try:
            wb = openpyxl.load_workbook(result_file)
            ws = wb.active
        except Exception as e:
            logging.exception(f"Erro ao abrir arquivo Excel existente: {result_file}")
            messagebox.showerror("Erro", f"Não foi possível abrir o arquivo Excel: {e}", icon="error")
            return

    pdf_count = 0
    error_messages = []
    linhas_digitaveis_processadas = set()
    total_lines = 0

    temp_dir_obj = tempfile.TemporaryDirectory()
    temp_dir = temp_dir_obj.name

    try:
        for pdf_path in input_files:
            n_processo = os.path.basename(pdf_path)
            if not n_processo.lower().endswith('.pdf'):
                continue

            pdf_count += 1
            popup_file_label.config(text=f"Processando: {n_processo}")
            root.update_idletasks()

            if not processing:
                break

            ocr_text, num_pages = ocr_pdf(pdf_path, temp_dir)
            if not processing:
                break

            if ocr_text:
                info = extract_info(ocr_text)
                nome_sem_extensao = os.path.splitext(n_processo)[0]
                total_lines += 1

                if not any(info.values()):
                    arquivos_com_dados_incompletos.add(nome_sem_extensao)
                    ws.append([nome_sem_extensao, '', '', '', f"Custas{custas}:{total_lines:02}"])
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        cell.font = Font(color='000000')
                    error_messages.append(f"Arquivo {n_processo}: Nenhuma informação encontrada.")
                    logging.warning(f"Arquivo {n_processo}: Nenhuma informação encontrada.")
                elif info['tipo'] == 'guia_custas' and info['cnpj'] != 'N/A':
                    arquivos_com_dados_incompletos.add(nome_sem_extensao)
                    ws.append([nome_sem_extensao, info['cnpj'], '', info['valor'] if info['valor'] else '', f"Custas{custas}:{total_lines:02}"])
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        cell.font = Font(color='000000')
                    error_messages.append(f"Arquivo {n_processo}: CNPJ encontrado, mas sem linha digitável (Guia de Custas?).")
                    logging.warning(f"Arquivo {n_processo}: CNPJ encontrado, mas sem linha digitável (Guia de Custas?).")
                elif info['cnpj'] == 'N/A':
                    arquivos_com_dados_incompletos.add(nome_sem_extensao)
                    num_linhas = len(info['linhas_digitaveis'])
                    if num_linhas > 0:
                        for i in range(num_linhas):
                            linha_digitavel = info['linhas_digitaveis'][i]
                            valor_monetario = info['valores'][i]
                            if linha_digitavel in linhas_digitaveis_processadas:
                                continue
                            linhas_digitaveis_processadas.add(linha_digitavel)
                            try:
                                valor_float = float(valor_monetario.replace(',', '.'))
                                valor_formatado = "{:,.2f}".format(valor_float).replace(',', '*').replace('.', ',').replace('*', '.')
                            except ValueError:
                                valor_formatado = valor_monetario
                            if i == 0:
                                ws.append([nome_sem_extensao, '', linha_digitavel, valor_formatado, f"Custas{custas}:{total_lines:02}"])
                            else:
                                ws.append([f"{nome_sem_extensao} - Boleto página {i + 1}", '', linha_digitavel, valor_formatado, f"Custas{custas}:{total_lines:02}"])
                    else:
                        ws.append([nome_sem_extensao, '', '', '', f"Custas{custas}:{total_lines:02}"])
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        cell.font = Font(color='000000')
                    error_messages.append(f"Arquivo {n_processo}: CNPJ não encontrado ou ignorado.")
                    logging.warning(f"Arquivo {n_processo}: CNPJ não encontrado ou ignorado.")
                else:
                    num_linhas = len(info['linhas_digitaveis'])
                    for i in range(num_linhas):
                        linha_digitavel = info['linhas_digitaveis'][i]
                        valor_monetario = info['valores'][i]
                        if linha_digitavel in linhas_digitaveis_processadas:
                            continue
                        linhas_digitaveis_processadas.add(linha_digitavel)
                        try:
                            valor_float = float(valor_monetario.replace(',', '.'))
                            valor_formatado = "{:,.2f}".format(valor_float).replace(',', '*').replace('.', ',').replace('*', '.')
                        except ValueError:
                            valor_formatado = valor_monetario
                        if i == 0:
                            ws.append([nome_sem_extensao, info['cnpj'], linha_digitavel, valor_formatado, f"Custas{custas}:{total_lines:02}"])
                        else:
                            ws.append([f"{nome_sem_extensao} - Boleto página {i + 1}", info['cnpj'], linha_digitavel, valor_formatado, f"Custas{custas}:{total_lines:02}"])
                        if 'N/A' in [v for k, v in info.items() if k != 'cnpj']:
                            error_messages.append(f"Arquivo {n_processo}: Dados inválidos ou ausentes (exceto CNPJ).")
                            arquivos_com_dados_incompletos.add(nome_sem_extensao)
                            logging.warning(f"Arquivo {n_processo}: Dados inválidos ou ausentes (exceto CNPJ).")

                for col in range(1, ws.max_column + 1):
                    column_letter = openpyxl.utils.get_column_letter(col)
                    column_width = max(len(str(cell.value)) if cell.value else 0 for cell in ws[column_letter])
                    ws.column_dimensions[column_letter].width = max(column_width, 10)

                if num_pages > 1:
                    arquivos_com_paginas_a_mais.add(nome_sem_extensao)
                    logging.warning(f"Arquivo {n_processo}: Possui mais de uma página.")
            else:
                error_messages.append(f"Arquivo {n_processo}: Falha no processamento do OCR.")
                arquivos_com_dados_incompletos.add(nome_sem_extensao)
                ws.append([nome_sem_extensao, '', '', '', f"Custas{custas}:{total_lines:02}"])
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.font = Font(color='000000')
                logging.error(f"Arquivo {n_processo}: Falha no processamento do OCR.")

            if processing:
                popup_progress_bar['value'] += 1
                root.update_idletasks()
            if not processing:
                break

        for row in range(2, ws.max_row + 1):
            nome_arquivo_celula = ws.cell(row, 1).value
            # Garante que nome_arquivo_celula é uma string antes de verificar a substring
            if isinstance(nome_arquivo_celula, str):
                 # Pega o nome base do arquivo, removendo sufixos como " - Boleto página X"
                nome_base_arquivo = nome_arquivo_celula.split(" - Boleto página ")[0]
                if nome_base_arquivo in arquivos_com_dados_incompletos:
                    for cell in ws[row]:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        cell.font = Font(color='000000')

        for cell in ws['D']:
            cell.number_format = 'R$ #,##0.00'
            cell.alignment = Alignment(horizontal='right')
        ws['D1'].alignment = Alignment(horizontal='left')

        for row in range(2, ws.max_row + 1):
            valor_boleto = ws.cell(row, 4).value
            if valor_boleto and isinstance(valor_boleto, str):
                try: # Adiciona try-except para valores não numéricos
                    valor_boleto = float(valor_boleto.replace('.', '').replace(',', '.'))
                except ValueError:
                    valor_boleto = None # Define como None se não puder converter
            if valor_boleto and valor_boleto > 2000:
                error_messages.append(f"Arquivo {ws.cell(row, 1).value}: Valor do boleto acima de R$ 2000. Verificar manual.")
                logging.warning(f"Arquivo {ws.cell(row, 1).value}: Valor do boleto acima de R$ 2000. Verificar manual.")
                for cell in ws[row]:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.font = Font(color='000000')
    except Exception as e:
        logging.exception("Erro durante o processamento dos PDFs")
        messagebox.showerror("Erro", f"Erro durante o processamento: {e}", icon="error")
    finally:
        if processing:
            try:
                wb.save(result_file)
                if save_csv and not error_messages and not arquivos_com_paginas_a_mais and not arquivos_com_dados_incompletos:
                    save_to_csv(result_file, ws)
            except Exception as e:
                logging.exception(f"Erro ao salvar o arquivo Excel: {result_file}")
                messagebox.showerror("Erro", f"Erro ao salvar o arquivo Excel: {e}", icon="error")
        temp_dir_obj.cleanup()

    arquivos_com_paginas_a_mais_list = list(arquivos_com_paginas_a_mais)
    arquivos_com_dados_incompletos_list = list(arquivos_com_dados_incompletos)

    if processing:
        if error_messages or arquivos_com_paginas_a_mais_list or arquivos_com_dados_incompletos_list:
            show_divergencia_popup(error_messages, result_file, arquivos_com_paginas_a_mais_list, arquivos_com_dados_incompletos_list, save_csv)
        else:
            show_success_popup(result_file, save_csv)
    else:
        show_cancelled_popup()

    return arquivos_com_paginas_a_mais, arquivos_com_dados_incompletos

def select_input_files():
    global input_files, input_dir
    input_files_tuple = filedialog.askopenfilenames( # Use um nome diferente para a tupla
        filetypes=[("Arquivos PDF", "*.pdf")],
        title="Selecione os arquivos PDF"
    )
    if input_files_tuple:
        input_files = list(input_files_tuple) # Converta para lista
        input_dir = os.path.dirname(input_files[0])
        input_dir_label.config(text=f"{len(input_files)} arquivos selecionados")
    else:
        input_files = [] # Garante que input_files seja uma lista vazia
        input_dir = ""
        input_dir_label.config(text="PDF não selecionado")


def select_result_file():
    global result_file
    result_file = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Planilhas Excel", "*.xlsx")],
        title="Salve a planilha de resultados",
    )
    result_file_label.config(text=f"{result_file}")

def process_pdfs_thread(input_files_list, output_dir_str, result_file_str, custas_str, save_csv_var_bool):
    global processing, popup_progress_bar

    if not isinstance(output_dir_str, str):
        logging.error(f"Caminho de saída inválido: {output_dir_str}")
        messagebox.showerror("Erro", "O caminho de saída (output_dir) não é uma string válida.", icon="error")
        processing = False
        start_button['state'] = NORMAL
        return

    processing = True
    try:
        os.environ["PATH"] = poppler_path + os.path.sep + os.environ["PATH"]
        popup_progress_bar['value'] = 0
        num_pdfs = len(input_files_list)
        popup_progress_bar['maximum'] = num_pdfs
        if num_pdfs == 0:
            messagebox.showinfo("Aviso", "Nenhum arquivo PDF selecionado.", icon="warning")
            processing = False
            if popup and popup.winfo_exists(): popup.destroy() # Fecha popup se existir
            start_button['state'] = NORMAL
            return

        save_csv_bool = save_csv_var_bool.get()
        process_pdfs(input_files_list, output_dir_str, result_file_str, custas_str, save_csv_bool)
        processing = False

    except Exception as e:
        logging.exception("Erro no thread de processamento")
        messagebox.showerror("Erro", f"Erro durante o processamento: {e}", icon="error")
        processing = False
    finally:
        processing = False
        if popup and popup.winfo_exists():
            popup.destroy()
        start_button['state'] = NORMAL
        logging.info("Processamento de PDFs finalizado.")

def create_button(parent, text, command, is_default=False, fg_color="#FF0000"):
    button = Button(parent, text=text, command=command,
                     font=("Segoe UI Bold", 10),
                     bg=parent.cget("bg"),
                     fg=fg_color,
                     relief=FLAT,
                     borderwidth=1,
                     padx=10,
                     pady=5)
    if is_default:
        button.bind("<Return>", lambda event: command())
        button.focus_set()
    return button

def show_divergencia_popup(error_messages, result_file, arquivos_com_paginas_a_mais, arquivos_com_dados_incompletos, save_csv):
    global processing, input_files # Adiciona input_files como global para ser acessível
    popup_div = Toplevel(root)
    popup_div.title("Processamento Concluído com Divergências")
    popup_div.transient(root)
    popup_div.grab_set()
    popup_div.resizable(False, False)

    label = Label(popup_div, text="Processamento concluído! Divergências encontradas:", font=("Segoe UI Bold", 10), fg="#FF0000")
    label.pack(pady=10)

    if arquivos_com_paginas_a_mais:
        label_paginas_a_mais = Label(popup_div, text=f"Arquivos com mais de uma página: {len(arquivos_com_paginas_a_mais)}", font=("Segoe UI Bold", 10))
        label_paginas_a_mais.pack(pady=5)
        # Criar um frame para a lista com scrollbar se necessário
        frame_paginas = Frame(popup_div)
        scrollbar_paginas = Scrollbar(frame_paginas, orient=VERTICAL)
        lista_paginas_text = Text(frame_paginas, wrap=WORD, height=min(5, len(arquivos_com_paginas_a_mais)), width=50, yscrollcommand=scrollbar_paginas.set, font=("Segoe UI", 9))
        for item in arquivos_com_paginas_a_mais:
            lista_paginas_text.insert(END, item + "\n")
        lista_paginas_text.config(state=DISABLED) # Torna o texto não editável
        scrollbar_paginas.config(command=lista_paginas_text.yview)
        scrollbar_paginas.pack(side=RIGHT, fill=Y)
        lista_paginas_text.pack(side=LEFT, fill=BOTH, expand=True)
        frame_paginas.pack(pady=5, padx=10, fill=X)


    if arquivos_com_dados_incompletos:
        label_dados_incompletos = Label(popup_div, text=f"Arquivos com informações faltando: {len(arquivos_com_dados_incompletos)}", font=("Segoe UI Bold", 10))
        label_dados_incompletos.pack(pady=5)
        # Criar um frame para a lista com scrollbar
        frame_dados = Frame(popup_div)
        scrollbar_dados = Scrollbar(frame_dados, orient=VERTICAL)
        lista_dados_text = Text(frame_dados, wrap=WORD, height=min(5, len(arquivos_com_dados_incompletos)), width=50, yscrollcommand=scrollbar_dados.set, font=("Segoe UI", 9))
        for item in arquivos_com_dados_incompletos:
            lista_dados_text.insert(END, item + "\n")
        lista_dados_text.config(state=DISABLED) # Torna o texto não editável
        scrollbar_dados.config(command=lista_dados_text.yview)
        scrollbar_dados.pack(side=RIGHT, fill=Y)
        lista_dados_text.pack(side=LEFT, fill=BOTH, expand=True)
        frame_dados.pack(pady=5, padx=10, fill=X)


    if error_messages: # Mostra mensagens de erro gerais se houver
        label_erros_gerais = Label(popup_div, text="Outras observações:", font=("Segoe UI Bold", 10))
        label_erros_gerais.pack(pady=5)
        frame_erros = Frame(popup_div)
        scrollbar_erros = Scrollbar(frame_erros, orient=VERTICAL)
        lista_erros_text = Text(frame_erros, wrap=WORD, height=min(5, len(error_messages)), width=50, yscrollcommand=scrollbar_erros.set, font=("Segoe UI", 9))
        for msg in error_messages:
            lista_erros_text.insert(END, msg + "\n")
        lista_erros_text.config(state=DISABLED)
        scrollbar_erros.config(command=lista_erros_text.yview)
        scrollbar_erros.pack(side=RIGHT, fill=Y)
        lista_erros_text.pack(side=LEFT, fill=BOTH, expand=True)
        frame_erros.pack(pady=5, padx=10, fill=X)


    if save_csv and (error_messages or arquivos_com_paginas_a_mais or arquivos_com_dados_incompletos):
        label_csv_nao_criado = Label(popup_div, text="(CSV automático não criado, por falta de confiabilidade)", font=("Segoe UI Bold", 10), fg="#FF0000")
        label_csv_nao_criado.pack(pady=5)

    button_frame = Frame(popup_div)
    button_frame.pack(pady=10)

    ok_button = create_button(button_frame, "OK", lambda: [popup_div.destroy(), webbrowser.open_new_tab(f"file://{result_file}")], is_default=True, fg_color="#000000")
    ok_button.pack(side=LEFT, padx=10)

    # Combina as duas listas para o botão "Conferir PDFs"
    todos_arquivos_para_conferir = list(set(arquivos_com_paginas_a_mais + arquivos_com_dados_incompletos))

    conferir_pdfs_button = Button(button_frame, text="Conferir PDFs",
                                  command=lambda: [popup_div.destroy(), abrir_arquivos_pdf(todos_arquivos_para_conferir), webbrowser.open_new_tab(f"file://{result_file}")],
                                  font=("Segoe UI Bold", 10), bg=popup_div.cget("bg"), fg="#0000FF", relief=FLAT, borderwidth=1, padx=10, pady=5)
    conferir_pdfs_button.pack(side=LEFT, padx=10)

    center_window(popup_div)


def show_success_popup(result_file, save_csv):
    popup_success = Toplevel(root)
    popup_success.title("Processamento Concluído com Sucesso")
    popup_success.transient(root)
    popup_success.grab_set()
    popup_success.resizable(False, False)

    label = Label(popup_success, text="Processamento concluído com sucesso!", font=("Segoe UI Bold", 10), fg="green")
    label.pack(pady=10)

    if save_csv:
        label_csv_criado = Label(popup_success, text="(CSV automático também foi criado)", font=("Segoe UI Bold", 10), fg="green")
        label_csv_criado.pack(pady=5)

    button_frame = Frame(popup_success)
    button_frame.pack(pady=10)

    ok_button = create_button(button_frame, "OK", lambda: [popup_success.destroy(), webbrowser.open_new_tab(f"file://{result_file}")], is_default=True, fg_color="#000000")
    ok_button.pack(side=LEFT, padx=10)

    center_window(popup_success)

def show_cancelled_popup():
    popup_cancel = Toplevel(root)
    popup_cancel.title("Processamento Cancelado!")
    popup_cancel.transient(root)
    popup_cancel.grab_set()
    popup_cancel.resizable(False, False)

    label = Label(popup_cancel, text="Processamento Cancelado!", font=("Segoe UI Bold", 10), fg="red")
    label.pack(pady=10)

    button_frame = Frame(popup_cancel)
    button_frame.pack(pady=10)

    ok_button = create_button(button_frame, "OK", popup_cancel.destroy, is_default=True, fg_color="#FF0000")
    ok_button.pack(side=LEFT, padx=10)

    center_window(popup_cancel)

def abrir_arquivos_pdf(arquivos_nomes_base):
    global input_files # Precisa acessar a lista de caminhos completos
    arquivos_abertos = 0
    if not input_files:
        logging.warning("Tentativa de abrir PDFs sem input_files definidos.")
        return

    for nome_base in arquivos_nomes_base:
        caminho_completo = next((f for f in input_files if os.path.splitext(os.path.basename(f))[0] == nome_base), None)
        if caminho_completo:
            try:
                webbrowser.open_new_tab(f"file://{caminho_completo}")
                arquivos_abertos +=1
            except Exception as e:
                logging.error(f"Erro ao tentar abrir PDF {caminho_completo}: {e}")
        else:
            logging.warning(f"Arquivo PDF com nome base '{nome_base}' não encontrado na lista de input_files.")
    if arquivos_abertos == 0 and arquivos_nomes_base:
        messagebox.showwarning("Aviso", "Não foi possível localizar os arquivos PDF para abrir. Verifique se foram movidos ou renomeados.", icon="warning")


def start_processing():
    global processing, popup, popup_file_label, popup_progress_bar, custas_entry, input_files, result_file, save_csv_var, input_dir

    if start_button['state'] == DISABLED:
        return

    processing = True # Define processing como True no início
    logging.info("Iniciando o processamento...")

    try:
        if not input_files: # Verifica se a lista input_files está vazia ou não definida
            raise NameError("input_files não definidos ou vazios")
    except NameError:
        logging.error("Arquivos PDF não selecionados.")
        messagebox.showerror("Erro", "Arquivos não Selecionados: Selecione os arquivos PDF.", icon="error")
        processing = False # Garante que processing seja False
        return

    try:
        result_file
        if not result_file: raise NameError # Verifica se está vazio
    except NameError:
        logging.error("Planilha de resultado não selecionada.")
        messagebox.showerror("Erro", "Planilha não Selecionada: Selecione a planilha de resultados.", icon="error")
        processing = False # Garante que processing seja False
        return


    if not os.path.exists(poppler_path):
        logging.error(f"Pasta do Poppler não encontrada: {poppler_path}")
        messagebox.showerror("Erro", f"Pasta do Poppler não encontrada em: {poppler_path}. O programa não poderá funcionar corretamente", icon="error")
        processing = False # Garante que processing seja False
        return

    custas = custas_entry.get()
    if not re.match(r'^[0-9\.\:\/\\]{0,5}$', custas): # Permite de 0 a 5 caracteres
        logging.error(f"Valor de custas inválido: {custas}")
        messagebox.showerror("Erro", "Digite um valor válido para as custas (apenas números, '.', ':', '/', '\\) com até 5 caracteres.", icon="error")
        processing = False # Garante que processing seja False
        return

    start_button['state'] = DISABLED

    popup = Toplevel(root)
    popup.title("Processando PDFs")
    popup.transient(root)
    popup.grab_set()
    popup.resizable(False, False)
    popup_file_label = Label(popup, text="Processando arquivo...", font=("Segoe UI Bold", 10))
    popup_file_label.pack(pady=10)
    popup_progress_bar = ttk.Progressbar(popup, orient="horizontal", length=300, mode="determinate")
    popup_progress_bar.pack(pady=10)
    cancel_button = Button(popup, text="Cancelar", command=cancel_processing,
                           font=("Segoe UI Bold", 10),
                           bg=popup.cget("bg"),
                           fg="#FF0000",
                           relief=FLAT,
                           borderwidth=1,
                           padx=10,
                           pady=5)
    cancel_button.pack(pady=10)
    center_window(popup)

    output_dir = input_dir if input_dir else os.getcwd()
    logging.info(f"Pasta de saída para arquivos temporários: {output_dir}")

    thread = Thread(target=process_pdfs_thread, args=(input_files, output_dir, result_file, custas, save_csv_var))
    thread.start()

def cancel_processing():
    global processing, popup
    processing = False
    logging.info("Processamento cancelado pelo usuário.")
    if popup and popup.winfo_exists():
        popup.destroy()
    show_cancelled_popup()
    start_button['state'] = NORMAL

def toggle_csv_save():
    save_csv_var.set(not save_csv_var.get())

# Função para criar botões arredondados (versão original para o botão 'i')
def create_rounded_button(parent, text, command, width=30, height=30, bg_color="#007bff", text_color="#FFFFFF"):
    canvas = Canvas(parent, width=width, height=height, bd=0, highlightthickness=0, relief='ridge', bg=parent.cget("bg"))
    # Desenha o círculo/oval
    # As coordenadas são (x1, y1, x2, y2) para o retângulo que circunscreve o oval
    # Adiciona uma pequena margem para a borda não ser cortada
    oval_id = canvas.create_oval(2, 2, width-2, height-2, outline=bg_color, fill=bg_color)
    # Adiciona o texto no centro
    text_id = canvas.create_text(width/2, height/2, text=text, fill=text_color, font=("Segoe UI Bold", int(height/2.5)))
    canvas.bind("<Button-1>", lambda event: command())
    return canvas

def open_log_file():
    try:
        webbrowser.open_new_tab(f"file://{log_file_path}")
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível abrir o arquivo de log: {e}", icon="error")

def delete_log_file():
    try:
        logging.shutdown()
        os.remove(log_file_path)
        logging.basicConfig(filename=log_file_path,
                            level=logging.INFO,
                            filemode='w',
                            format='%(asctime)s - %(levelname)s - %(message)s')
        logging.info("Arquivo de log deletado pelo usuário.")
        messagebox.showinfo("Sucesso", "Arquivo de log deletado com sucesso.", icon="info")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo de log não encontrado para deletar.", icon="error")
    except PermissionError:
        messagebox.showerror("Erro", "Não foi possível deletar o arquivo de log. Ele pode estar em uso ou você não tem permissão.", icon="error")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao deletar o arquivo de log: {e}", icon="error")

def show_filtro_cnpj_config_popup(parent_window):
    global ignored_cnpjs_list # Acessa a lista global

    config_popup = Toplevel(parent_window)
    config_popup.title("Configurar Filtro de CNPJ")
    config_popup.transient(parent_window)
    config_popup.grab_set()
    config_popup.resizable(False, False)
    config_popup.configure(bg="#f0f0f0") # Cor de fundo similar a outros popups

    main_frame = Frame(config_popup, padx=15, pady=15, bg=config_popup.cget("bg"))
    main_frame.pack(fill=BOTH, expand=True)

    Label(main_frame, text="CNPJs a serem ignorados (separados por vírgula):", font=("Segoe UI", 10), bg=main_frame.cget("bg")).pack(pady=(0,5), anchor="w")

    cnpj_entry_var = StringVar()
    # Carrega os CNPJs atuais no campo de entrada
    cnpj_entry_var.set(",".join(ignored_cnpjs_list))

    cnpj_entry = Entry(main_frame, textvariable=cnpj_entry_var, width=60, font=("Segoe UI", 10))
    cnpj_entry.pack(pady=5, fill=X)

    button_frame = Frame(main_frame, bg=main_frame.cget("bg"))
    button_frame.pack(pady=(10,0), fill=X)

    def on_save():
        cnpjs_text = cnpj_entry_var.get()
        # Validação simples de formato (opcional, mas útil)
        cnpjs_to_save = []
        has_invalid = False
        for c in cnpjs_text.split(','):
            c = c.strip()
            if c: # Adiciona apenas se não for vazio
                if re.match(r'^\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}$', c):
                    cnpjs_to_save.append(c)
                else:
                    has_invalid = True
        
        if has_invalid:
            if not messagebox.askyesno("CNPJ Inválido", "Um ou mais CNPJs parecem estar em formato inválido. Deseja salvar mesmo assim?", icon="warning", parent=config_popup):
                return # Não salva se o usuário cancelar


        final_cnpjs_string = ",".join(cnpjs_to_save)
        if save_cnpjs_to_config(final_cnpjs_string): # save_cnpjs_to_config já recarrega a lista
            messagebox.showinfo("Sucesso", "Filtro de CNPJs salvo com sucesso!", parent=config_popup)
            config_popup.destroy()
        # Não precisa de else, save_cnpjs_to_config já mostra erro

    def on_open_config_file():
        try:
            if not os.path.exists(filtro_config_path):
                # Cria o arquivo com o conteúdo atual do entry se não existir
                save_cnpjs_to_config(cnpj_entry_var.get())
            
            # Tenta abrir com o programa padrão do sistema
            if sys.platform == "win32":
                os.startfile(filtro_config_path)
            elif sys.platform == "darwin": # macOS
                subprocess.call(["open", filtro_config_path])
            else: # linux variants
                subprocess.call(["xdg-open", filtro_config_path])
            messagebox.showinfo("Informação", f"Após editar e salvar o arquivo '{os.path.basename(filtro_config_path)}'(Caso tenha feito fora do Editor do programa), clique em 'cancelar' nesta janela para aplicar as mudanças", parent=config_popup, icon="info")

        except Exception as e:
            logging.error(f"Erro ao abrir arquivo de configuração de CNPJ: {e}")
            messagebox.showerror("Erro", f"Não foi possível abrir o arquivo de configuração: {e}", icon="error", parent=config_popup)


    save_button = Button(button_frame, text="Salvar", command=on_save, font=("Segoe UI Bold", 10), bg="#4CAF50", fg="white", relief=FLAT, padx=10, pady=5)
    save_button.pack(side=LEFT, padx=5)

    open_config_button = Button(button_frame, text="Abrir Filtro.config", command=on_open_config_file, font=("Segoe UI Bold", 10), bg="#2196F3", fg="white", relief=FLAT, padx=10, pady=5)
    open_config_button.pack(side=LEFT, padx=5)

    cancel_button = Button(button_frame, text="Cancelar", command=config_popup.destroy, font=("Segoe UI Bold", 10), bg="#F44336", fg="white", relief=FLAT, padx=10, pady=5)
    cancel_button.pack(side=RIGHT, padx=5) # Alinha à direita

    center_window(config_popup)
    config_popup.focus_set() # Garante que o popup tenha foco
    cnpj_entry.focus() # Coloca foco no campo de entrada


root = Tk()
root.title("Extrator de Dados Do Boleto (1.3.6a)") # Versão atualizada
font_style = font.Font(family="Segoe UI Bold", size=15)
root.option_add("*Font", font_style)
root.resizable(False, False)
frame = Frame(root)
frame.pack(padx=20, pady=20)

input_dir_button = Button(frame, text=" Selecionar PDFs ", command=select_input_files)
input_dir_button.grid(row=0, column=0, padx=5, pady=5, sticky="e")
input_dir_label = Label(frame, text="PDF não selecionado")
input_dir_label.grid(row=0, column=1, padx=5, pady=5, sticky="w")

result_file_button = Button(frame, text="Local do Resultado", command=select_result_file)
result_file_button.grid(row=1, column=0, padx=5, pady=5, sticky="e")
result_file_label = Label(frame, text="Arquivo não selecionado")
result_file_label.grid(row=1, column=1, padx=5, pady=5, sticky="w")

save_csv_var = BooleanVar()
save_csv_label = Label(frame, text="CSV ponto e vírgula:", cursor="hand2")
save_csv_label.grid(row=2, column=0, pady=5, sticky="e")
save_csv_label.bind("<Button-1>", lambda event: toggle_csv_save())
save_csv_check = Checkbutton(frame, variable=save_csv_var, command=toggle_csv_save)
save_csv_check.grid(row=2, column=1, pady=5, sticky="w")

custas_label = Label(frame, text="Custas:", anchor='e', justify='left')
custas_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
custas_entry = Entry(frame, width=10)
custas_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

def limit_custas_entry(event):
    # Permite que o campo fique vazio
    current_text = custas_entry.get()
    if len(current_text) > 5:
        # Se exceder 5, corta para 5.
        # Precisamos de um pequeno truque para evitar recursão com <KeyRelease>
        # root.after_idle(lambda: custas_entry.delete(5, END)) # Isso pode causar problemas com a digitação rápida
        new_text = current_text[:5]
        if custas_entry.get() != new_text: # Evita loop infinito
            custas_entry.delete(0, END)
            custas_entry.insert(0, new_text)
custas_entry.bind("<KeyRelease>", limit_custas_entry)


# Botão de Iniciar processamento (estilo original)
start_button = Button(root, text="             Iniciar Processamento             ", command=start_processing,
                           font=("Segoe UI Bold", 15), # Fonte um pouco maior
                           bg="#4CAF50",  # Verde
                           fg="white",    # Texto branco
                           relief=FLAT,   # Sem borda 3D pronunciada
                           padx=10,       # Padding horizontal
                           pady=5)        # Padding vertical
start_button.pack(pady=10)


def show_info():
    info_popup = Toplevel(root)
    info_popup.title("Informação")
    info_popup.transient(root)
    info_popup.grab_set()
    info_popup.resizable(False, False)
    info_popup.configure(bg="#f0f0f0") # Cor de fundo padrão

    # Frame principal para conteúdo
    content_frame = Frame(info_popup, padx=15, pady=15, bg=info_popup.cget("bg"))
    content_frame.pack(expand=True, fill=BOTH)

    version_label = Label(content_frame, text=f"{root.title()} - by Elias", font=("Segoe UI", 10), bg=content_frame.cget("bg"), fg="#002b00")
    version_label.pack(pady=(0,5))
    pix_label = Label(content_frame, text="Chamado via mensagem Pix: eliasgkersten@gmail.com", font=("Segoe UI", 10), bg=content_frame.cget("bg"), fg="#002b00")
    pix_label.pack(pady=5)

    button_frame = Frame(content_frame, bg=content_frame.cget("bg"))
    button_frame.pack(pady=(10,0))

    # Botão Debug (verde)
    debug_button = Button(button_frame, text="Debug", command=open_log_file, font=("Segoe UI Bold", 10), bg="#4CAF50", fg="white", relief=FLAT, padx=10, pady=5)
    debug_button.pack(side=LEFT, padx=5)

    # Botão C Debug (azul)
    c_debug_button = Button(button_frame, text="🗑️Excluir Debug", command=delete_log_file, font=("Segoe UI Bold", 10), bg="#2196F3", fg="white", relief=FLAT, padx=10, pady=5)
    c_debug_button.pack(side=LEFT, padx=5)

    # Botão Configurações Filtro CNPJ (ícone de engrenagem)
    config_button = Button(button_frame, text="🛠️ Filtro", command=lambda: show_filtro_cnpj_config_popup(info_popup), font=("Segoe UI Bold", 10), bg="#607D8B", fg="white", relief=FLAT, padx=10, pady=5) # Cinza azulado
    config_button.pack(side=LEFT, padx=5)

    # Botão Sair (vermelho)
    exit_button = Button(button_frame, text="Sair", command=info_popup.destroy, font=("Segoe UI Bold", 10), bg="#F44336", fg="white", relief=FLAT, padx=10, pady=5)
    exit_button.pack(side=LEFT, padx=5)

    center_window(info_popup)


# Botão de Informação "i" (estilo original azul redondo)
show_info_button_canvas = create_rounded_button(root, "i", show_info, width=30, height=30, bg_color="#007bff", text_color="#FFFFFF")
# Posicionamento do botão de informação no canto inferior direito
show_info_button_canvas.place(relx=1.0, rely=1.0, x=-10, y=-10, anchor="se") # x e y negativos para dar uma margem da borda


center_window(root)

root.bind("<Return>", lambda event: start_processing() if not root.grab_current() else "break")

processing = False
input_files = [] # Inicializa input_files como uma lista vazia
result_file = "" # Inicializa result_file
input_dir = ""

if getattr(sys, 'frozen', False):
    # Tenta ocultar o console de forma mais robusta
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    # subprocess.run('cmd /c "exit"', shell=True, startupinfo=startupinfo) # Isso apenas executa e fecha um cmd
    # A melhor forma de não ter console é compilar com `pyinstaller --noconsole seu_script.py`

root.mainloop()